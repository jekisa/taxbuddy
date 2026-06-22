"""
Pajak Keluaran Otomatis - Web Edition
Flask backend porting the original Tkinter desktop logic.
"""

import os, re, io, json, uuid, shutil
from datetime import datetime, date
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from collections import defaultdict, Counter
from xml.etree.ElementTree import Element, SubElement, tostring
from xml.dom import minidom

from flask import Flask, render_template, request, jsonify, session, send_file

import xlrd, openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

_DIR = os.path.dirname(os.path.abspath(__file__))
_SOURCE_DATA_DIR = os.path.join(_DIR, "data")


def _resolve_data_dir():
    configured = os.environ.get("CORTEX_DATA_DIR") or os.environ.get("DATA_DIR")
    if configured:
        return configured
    if os.environ.get("VERCEL"):
        return os.path.join("/tmp", "cortex-data")
    return _SOURCE_DATA_DIR


DATA_DIR = _resolve_data_dir()
os.makedirs(DATA_DIR, exist_ok=True)


def _seed_data_dir():
    if os.path.abspath(DATA_DIR) == os.path.abspath(_SOURCE_DATA_DIR):
        return
    if not os.path.isdir(_SOURCE_DATA_DIR):
        return
    for name in os.listdir(_SOURCE_DATA_DIR):
        src = os.path.join(_SOURCE_DATA_DIR, name)
        dst = os.path.join(DATA_DIR, name)
        if os.path.isfile(src) and name.endswith(".json") and not os.path.exists(dst):
            shutil.copyfile(src, dst)


_seed_data_dir()
TEMPLATES_FILE  = os.path.join(DATA_DIR, "column_templates.json")
DLM_TEMPLATES_FILE = os.path.join(DATA_DIR, "dlm_column_templates.json")
DB_PENJUAL_FILE = os.path.join(DATA_DIR, "db_penjual.json")
DB_PEMBELI_FILE = os.path.join(DATA_DIR, "db_pembeli.json")
DB_PEMASOK_FILE = os.path.join(DATA_DIR, "db_pemasok.json")
SETTINGS_FILE   = os.path.join(DATA_DIR, "settings.json")
SECRET_KEY_FILE = os.path.join(DATA_DIR, ".secret_key")
EXPORT_HISTORY_FILE = os.path.join(DATA_DIR, "export_history.json")

TPK_FIELDS      = ["Nama Pelanggan","Keterangan Barang","No. Faktur",
                   "Tgl Faktur","Kuantitas","Harga Satuan","Unit 1 Barang"]
REQUIRED_FIELDS = set(TPK_FIELDS)

# ---- THEMES ----------------------------------------------------------------
THEMES = {
    "Dark Executive": {
        "bg":"#111827","surface":"#172033","surface2":"#1F2937","surface3":"#273449",
        "border":"#334155","border2":"#475569",
        "accent":"#38BDF8","accent2":"#2DD4BF","accent_dim":"#164E63",
        "gold":"#D6A84F","green":"#2F9E68","red":"#D65A5A","purple":"#8B7AD9",
        "text":"#F8FAFC","text2":"#CBD5E1","text3":"#8FA3B8",
        "row_a":"#182033","row_b":"#202A3D","row_sel":"#28415E","row_hdr":"#111827",
        "sidebar_w": 230,
    },
    "Midnight Blue": {
        "bg":"#0F172A","surface":"#142036","surface2":"#1E2B44","surface3":"#243653",
        "border":"#31415F","border2":"#4A5E7F",
        "accent":"#60A5FA","accent2":"#67E8F9","accent_dim":"#1E3A5F",
        "gold":"#D9B66B","green":"#45A67A","red":"#D86A6A","purple":"#A78BFA",
        "text":"#F1F5F9","text2":"#C3D0E3","text3":"#8191AA",
        "row_a":"#152039","row_b":"#1B2944","row_sel":"#243B63","row_hdr":"#0F172A",
        "sidebar_w": 230,
    },
    "Forest Dark": {
        "bg":"#101815","surface":"#16231F","surface2":"#1D2D28","surface3":"#263A33",
        "border":"#345047","border2":"#486A5F",
        "accent":"#66C28A","accent2":"#8DD3C7","accent_dim":"#1D4A34",
        "gold":"#D8B45D","green":"#66C28A","red":"#D46A64","purple":"#9B8FD9",
        "text":"#F2F7F4","text2":"#BDD3C7","text3":"#7B9589",
        "row_a":"#18231F","row_b":"#1F2D28","row_sel":"#254536","row_hdr":"#101815",
        "sidebar_w": 230,
    },
    "Light Pro": {
        "bg":"#F6F8FB","surface":"#FFFFFF","surface2":"#EEF3F8","surface3":"#E2E8F0",
        "border":"#D7DEE8","border2":"#B7C4D4",
        "accent":"#2563EB","accent2":"#0F766E","accent_dim":"#DCEBFA",
        "gold":"#B7791F","green":"#047857","red":"#B91C1C","purple":"#6D5BD0",
        "text":"#111827","text2":"#334155","text3":"#64748B",
        "row_a":"#FFFFFF","row_b":"#F8FAFC","row_sel":"#DBEAFE","row_hdr":"#EEF2F7",
        "sidebar_w": 230,
    },
    "Warm Latte": {
        "bg":"#F8F5F0","surface":"#FFFFFF","surface2":"#F1EBE2","surface3":"#E7DED2",
        "border":"#D6CABC","border2":"#B8A995",
        "accent":"#9A6A3A","accent2":"#3C8A7A","accent_dim":"#E9D8C7",
        "gold":"#A66F2D","green":"#3C8A62","red":"#B45345","purple":"#7A67B8",
        "text":"#2F241C","text2":"#645348","text3":"#958578",
        "row_a":"#FFFFFF","row_b":"#FAF7F2","row_sel":"#EDE2D6","row_hdr":"#F1EBE2",
        "sidebar_w": 230,
    },
    "Dracula": {
        "bg":"#171A22","surface":"#1F2430","surface2":"#2A3040","surface3":"#343B4E",
        "border":"#3F475C","border2":"#56627A",
        "accent":"#8EA7E9","accent2":"#75C9B7","accent_dim":"#2D3B5E",
        "gold":"#D8BC72","green":"#72B98F","red":"#D66F78","purple":"#A899D8",
        "text":"#F4F6FA","text2":"#C8D0DC","text3":"#8793A8",
        "row_a":"#202635","row_b":"#282E40","row_sel":"#38465F","row_hdr":"#171A22",
        "sidebar_w": 230,
    },
}

DUP_COLORS = [
    ("#FF6B8A","#3A0F1E"),("#FFB347","#3A2208"),("#56E0A0","#083A1A"),
    ("#B48EF7","#200A3A"),("#56C8E8","#082838"),("#FFD580","#3A2A08"),
    ("#FF8C69","#3A1508"),("#82CFFF","#081838"),
]

OUT_COLS = [
    ("No",52,"center"),("Nama Pelanggan",160,"left"),("Keterangan Barang",280,"left"),
    ("No. Faktur",130,"center"),("Tgl Faktur",100,"center"),
    ("Kuantitas",78,"right"),("Harga Satuan",100,"right"),("Unit",52,"center"),
    ("Harga Jual",118,"right"),("Diskon",108,"right"),("DPP",118,"right"),
    ("DPP Nilai Lain",128,"right"),("Tarif PPN",72,"center"),("PPN",118,"right"),
    ("Tarif PPnBM",82,"center"),("PPnBM",78,"right"),
]

FAKTUR_COLS = [
    ("Baris",52,"center"),("Tgl Faktur",100,"center"),
    ("Jenis Faktur",88,"center"),("Kode Transaksi",88,"center"),
    ("Ket. Tambahan",98,"center"),("Dok. Pendukung",98,"center"),
    ("Referensi",138,"center"),("Cap Fasilitas",88,"center"),
    ("ID TKU Penjual",168,"center"),("NPWP/NIK Pembeli",138,"center"),
    ("Jenis ID",68,"center"),("Negara",68,"center"),
    ("No. Dok Pembeli",168,"center"),("Nama Pembeli",198,"left"),
    ("Alamat Pembeli",278,"left"),("Email Pembeli",118,"center"),
    ("ID TKU Pembeli",168,"center"),
]

DET_COLS = [
    ("Baris",52,"center"),("Brg/Jasa",92,"center"),
    ("Kode",78,"center"),("Nama Barang / Jasa",298,"left"),
    ("Satuan",126,"center"),("Harga Satuan",98,"right"),
    ("Jumlah",78,"right"),("Total Diskon",108,"right"),
    ("DPP",118,"right"),("DPP Nilai Lain",128,"right"),
    ("Tarif PPN",72,"center"),("PPN",118,"right"),
    ("Tarif PPnBM",82,"center"),("PPnBM",78,"right"),
]

DETAIL_OPT_LABELS = {"A":"Barang", "B":"Jasa"}
DETAIL_UNIT_LABELS = {
    "UM.0003":"Kilogram", "UM.0004":"Gram", "UM.0013":"Meter",
    "UM.0014":"Inches", "UM.0015":"Sentimeter", "UM.0016":"Yard",
    "UM.0017":"Lusin", "UM.0018":"Unit", "UM.0019":"Set",
    "UM.0021":"Piece",
}
DETAIL_OPT_CHOICES = [f"{k} - {v}" for k,v in DETAIL_OPT_LABELS.items()]
DETAIL_UNIT_CHOICES = [f"{k} - {v}" for k,v in DETAIL_UNIT_LABELS.items()]

FIELD_ICONS = {
    "Nama Pelanggan":"person", "Keterangan Barang":"box", "No. Faktur":"hash",
    "Tgl Faktur":"calendar", "Kuantitas":"hash", "Harga Satuan":"tag", "Unit 1 Barang":"box",
}

# ---- DOC LAIN MASUKAN (input special documents, e.g. PPN PMSE TikTok) -------
DLM_FIELDS = ["No. Faktur", "Tgl Faktur", "Nama Pemasok", "Nilai Faktur"]
DLM_REQUIRED_FIELDS = set(DLM_FIELDS)

DLM_FIELD_ICONS = {
    "No. Faktur":"hash", "Tgl Faktur":"calendar",
    "Nama Pemasok":"person", "Nilai Faktur":"tag",
}

DLM_OUT_COLS = [
    ("No",50,"center"),("No. Dokumen",150,"left"),("Tgl Dokumen",100,"center"),
    ("Jenis Trx",70,"center"),("Kode Trx",70,"center"),("Jenis Dok",70,"center"),
    ("Masa",60,"center"),("Tahun",70,"center"),
    ("DPP (Tax Base)",118,"right"),("PPN (VAT)",110,"right"),("PPnBM (STLG)",100,"right"),
    ("NPWP Pemasok",150,"center"),("Nama Pemasok",180,"left"),
]

DLM_TRX_TYPE_LABELS = {
    "01":"DELIVERY",
    "02":"EXPORT",
    "03":"IMPORT",
    "04":"PURCHASE",
    "05":"UNCREDIT",
}

DLM_TRX_DETAIL_LABELS = {
    "01":"To other party than VAT Collector",
    "02":"To Government Institution as VAT Collector",
    "03":"To VAT Collector other than Government Institution",
    "04":"Other Tax Base",
    "05":"Certain value Tax Base",
    "06":"To Tourist for VAT Refund for Tourist",
    "07":"VAT Uncollected",
    "08":"VAT Exempted",
    "09":"Sale of unrelated to business asset",
    "10":"Other delivery of goods/services",
    "11":"Exports of Tangible Goods",
    "12":"Exports of Intangible Goods",
    "13":"Exports of Services",
    "14":"Import Taxable Goods",
    "15":"Utilization of Intangible Taxable Goods and Taxable Services",
}

DLM_TRX_DOC_LABELS = {
    "1":"Documents Equivalent to Tax Invoice",
    "2":"Cigarette Excise",
    "3":"Bonded Area Document",
    "4":"Export Notice of Goods",
    "5":"Export Notice of Services/Intangible Goods",
    "6":"Import Notice and Payment",
    "7":"Payment",
    "8":"Special documents",
    "9":"Notice of Tax Underpayment Assessment/ Notice of Additional Tax Underpayment Assessment",
    "10":"Import Notice",
}

DLM_TRX_TYPE_CHOICES = [f"{k} - {v}" for k,v in DLM_TRX_TYPE_LABELS.items()]
DLM_TRX_DETAIL_CHOICES = [f"{k} - {v}" for k,v in DLM_TRX_DETAIL_LABELS.items()]
DLM_TRX_DOC_CHOICES = [f"{k} - {v}" for k,v in DLM_TRX_DOC_LABELS.items()]

# ---- SPT DOKUMEN LAIN (pajak keluaran digunggung / RetailInvoiceBulk) --------
# Input manual: user mengisi nominal PPN (VAT), DPP dihitung mundur (rumus
# sheet "RUMUS PPN" template GUNGUNG): OtherTaxBase = VAT*100/12, TaxBase = Other/11*12.
SDL_OUT_COLS = [
    ("No",50,"center"),("TrxCode",80,"center"),("BuyerName",180,"left"),
    ("BuyerIdOpt",90,"center"),("BuyerIdNumber",150,"center"),
    ("GoodServiceOpt",100,"center"),("SerialNo",120,"left"),
    ("TransactionDate",110,"center"),
    ("TaxBaseSellingPrice",140,"right"),("OtherTaxBaseSellingPrice",160,"right"),
    ("VAT",120,"right"),("STLG",90,"right"),("Info",60,"center"),
]

SDL_TRX_CODE_LABELS = {
    "Normal":"Untuk Lampiran I.A.5",
    "07":"Untuk Lampiran I.A.9",
    "08":"Untuk Lampiran I.A.9",
    "NoVAT":"Untuk Lampiran I.B",
}

SDL_BUYER_ID_LABELS = {
    "NPWP":"Nomor Pokok Wajib Pajak",
    "NIK":"Nomor Induk Kependudukan",
}

SDL_GOOD_SERVICE_LABELS = DETAIL_OPT_LABELS

SDL_TRX_CODE_CHOICES = [f"{k} - {v}" for k,v in SDL_TRX_CODE_LABELS.items()]
SDL_BUYER_ID_CHOICES = [f"{k} - {v}" for k,v in SDL_BUYER_ID_LABELS.items()]
SDL_GOOD_SERVICE_CHOICES = DETAIL_OPT_CHOICES


def _choice_code(value,default):
    text=str(value or "").strip()
    if not text: return default
    return text.split(" - ",1)[0].strip()

def _detail_opt_label(code):
    code=(code or "A").strip()
    return f"{code} - {DETAIL_OPT_LABELS.get(code,code)}"

def _detail_unit_label(code):
    code=(code or "UM.0021").strip()
    return f"{code} - {DETAIL_UNIT_LABELS.get(code,code)}"


# ---- Helpers ----------------------------------------------------------------
def _jload(path):
    if os.path.exists(path):
        try:
            with open(path,"r",encoding="utf-8") as f: return json.load(f)
        except Exception: pass
    return {}

def _jsave(path,data):
    with open(path,"w",encoding="utf-8") as f:
        json.dump(data,f,ensure_ascii=False,indent=2)

def fmt(val,dec=2):
    if val in ("",None): return ""
    try: return f"{float(val):,.{dec}f}"
    except Exception: return str(val)

def fmt_idr(val):
    if val in ("",None,0,0.0): return "-"
    try: return f"{float(val):,.0f}"
    except Exception: return str(val)

def fmt_pct(val):
    if val in ("",None): return ""
    try:
        s=f"{round2(val):.2f}".rstrip("0").rstrip(".")
        return s
    except Exception:
        return str(val)

def round2(val):
    try:
        return float(Decimal(str(val)).quantize(Decimal("0.01"),rounding=ROUND_HALF_UP))
    except (InvalidOperation,ValueError,TypeError):
        return 0.0

def xml_num(val):
    return f"{round2(val):.2f}"

def xml_num_min(val):
    n=round2(val)
    if n==int(n): return str(int(n))
    return f"{n:.2f}".rstrip("0").rstrip(".")

def parse_num(val):
    if val in ("",None): return 0.0
    if isinstance(val,(int,float)): return float(val)
    s=str(val).strip()
    if not s: return 0.0
    s=re.sub(r"[^\d,.\-]","",s)
    if "," in s and "." in s:
        if s.rfind(",")>s.rfind("."):
            s=s.replace(".","").replace(",",".")
        else:
            s=s.replace(",","")
    elif "," in s:
        parts=s.split(",")
        s=("".join(parts[:-1])+"."+parts[-1]) if len(parts[-1])<=2 else s.replace(",","")
    elif "." in s:
        parts=s.split(".")
        if len(parts)>1 and all(len(p)==3 for p in parts[1:]):
            s="".join(parts)
    try: return float(s)
    except ValueError: return 0.0

def norm_party_name(val):
    return re.sub(r"[^a-z0-9]+","",str(val or "").casefold())

def apply_tax_calculation(row,diskon_pct=0):
    pct=max(parse_num(diskon_pct),0)
    hj=round2(row["qty"]*row["harga"])
    dk=round2(hj*pct/100)
    dpp=round2(hj-dk)
    nl=round2(dpp*(11/12))
    row.update({"harga_jual":hj,"diskon_rate":round2(pct),"diskon":dk,"dpp":dpp,
                "dpp_nl":nl,"tarif_ppn":"12","ppn":round2(nl*0.12),
                "tarif_ppnbm":"0","ppnbm":0})
    return row

def apply_dlm_calculation(row):
    tb=round2(row.get("nilai_faktur",0))
    row["tax_base"]=tb
    row["vat"]=round2(tb*11/111)
    row.setdefault("stlg",0.0)
    return row

def apply_sdl_calculation(row):
    vat=round2(row.get("vat",0))
    row["vat"]=vat
    other=round2(vat*100/12)
    row["other_tax_base"]=other
    row["tax_base"]=round2(other/11*12)
    row.setdefault("stlg",0.0)
    return row

def autofit_columns(ws):
    for col in ws.columns:
        mx=0; cl=get_column_letter(col[0].column)
        for cell in col:
            try: mx=max(mx,len(str(cell.value) if cell.value else ""))
            except Exception: pass
        ws.column_dimensions[cl].width=min(mx+4,80)

MONTH_NAME_TO_NUMBER = {
    "jan":"01", "january":"01", "januari":"01",
    "feb":"02", "february":"02", "februari":"02",
    "mar":"03", "march":"03", "maret":"03",
    "apr":"04", "april":"04",
    "may":"05", "mei":"05",
    "jun":"06", "june":"06", "juni":"06",
    "jul":"07", "july":"07", "juli":"07",
    "aug":"08", "august":"08", "agu":"08", "ags":"08", "agustus":"08",
    "sep":"09", "sept":"09", "september":"09",
    "oct":"10", "october":"10", "okt":"10", "oktober":"10",
    "nov":"11", "november":"11",
    "dec":"12", "december":"12", "des":"12", "desember":"12",
}

MONTH_LABELS_ID = {
    1:"Januari", 2:"Februari", 3:"Maret", 4:"April", 5:"Mei", 6:"Juni",
    7:"Juli", 8:"Agustus", 9:"September", 10:"Oktober", 11:"November", 12:"Desember",
}

def _date_to_iso(tgl):
    if not tgl: return ""
    if isinstance(tgl,(datetime,date)): return tgl.strftime("%Y-%m-%d")
    tgl=str(tgl).strip()
    if re.match(r"\d{4}-\d{2}-\d{2}",tgl):
        try: return datetime.strptime(tgl[:10],"%Y-%m-%d").strftime("%Y-%m-%d")
        except ValueError: return tgl[:10]
    m=re.match(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})$",tgl)
    if m:
        d,mn,y=m.groups()
        try: return datetime.strptime(f"{int(d):02d}/{int(mn):02d}/{y}","%d/%m/%Y").strftime("%Y-%m-%d")
        except ValueError: return tgl
    m=re.match(r"(\d{1,2})\s+([A-Za-z.]+)\s+(\d{4})$",tgl,re.IGNORECASE)
    if m:
        d,mn,y=m.groups()
        month=MONTH_NAME_TO_NUMBER.get(mn.rstrip(".").casefold())
        if not month:
            return tgl
        try:
            return datetime.strptime(f"{int(d):02d}/{month}/{y}","%d/%m/%Y").strftime("%Y-%m-%d")
        except ValueError:
            return tgl
    return tgl

def _date_to_ddmmyyyy(tgl):
    iso=_date_to_iso(tgl)
    m=re.match(r"(\d{4})-(\d{2})-(\d{2})$",iso or "")
    if m:
        y,mn,d=m.groups()
        return f"{d}/{mn}/{y}"
    return str(tgl or "").strip()

def _normalize_invoice_date(tgl):
    tgl=str(tgl or "").strip()
    if not tgl:
        return None
    normalized=_date_to_ddmmyyyy(tgl)
    if not _is_ddmmyyyy(normalized):
        return None
    return normalized

def _parse_ddmmyyyy_date(tgl):
    try:
        return datetime.strptime(str(tgl or "").strip(),"%d/%m/%Y").date()
    except ValueError:
        return None

def _is_ddmmyyyy(tgl):
    try:
        datetime.strptime(str(tgl or "").strip(),"%d/%m/%Y")
        return True
    except ValueError:
        return False

def _short_date_row(r):
    ref=str(r.get("referensi","")).strip() or f"baris {r.get('baris','')}".strip()
    tgl=str(r.get("tgl_faktur","")).strip() or "-"
    raw=str(r.get("tgl_raw","")).strip()
    if raw and raw!=tgl:
        return f"{ref}: {tgl} (sumber: {raw})"
    return f"{ref}: {tgl}"

def collect_faktur_date_diagnostics(faktur_rows):
    invalid=[]; valid=[]
    for i,r in enumerate(faktur_rows):
        parsed=_parse_ddmmyyyy_date(r.get("tgl_faktur",""))
        if parsed:
            valid.append((i,r,parsed))
        else:
            invalid.append((i,r))

    warnings=[]; blockers=[]; outliers=[]; majority_label=""
    if invalid:
        examples=", ".join(_short_date_row(r) for _i,r in invalid[:5])
        blockers.append(f"Tanggal Faktur kosong/tidak valid: {len(invalid)} baris. Contoh: {examples}")

    if valid:
        month_counts=Counter((d.year,d.month) for _i,_r,d in valid)
        (majority_year,majority_month),majority_count=month_counts.most_common(1)[0]
        majority_label=f"{MONTH_LABELS_ID[majority_month]} {majority_year}"
        if len(month_counts)>1 and majority_count>len(valid)/2:
            outliers=[(i,r,d) for i,r,d in valid if (d.year,d.month)!=(majority_year,majority_month)]
            if outliers:
                examples=", ".join(_short_date_row(r) for _i,r,_d in outliers[:8])
                warnings.append(
                    f"Tanggal Faktur di luar bulan mayoritas {majority_label}: "
                    f"{len(outliers)} baris. Contoh: {examples}"
                )

    invalid_indices=[i for i,_r in invalid]
    outlier_indices=[i for i,_r,_d in outliers]
    return {
        "majority_month": majority_label,
        "invalid_indices": invalid_indices,
        "outlier_indices": outlier_indices,
        "problem_indices": sorted(set(invalid_indices+outlier_indices)),
        "warnings": warnings,
        "blockers": blockers,
    }

def _cell_text(val):
    if val is None: return ""
    if isinstance(val,(datetime,date)): return val.strftime("%d/%m/%Y")
    if isinstance(val,float) and val.is_integer(): return str(int(val))
    return str(val).strip()


# ---- Data logic ---------------------------------------------------------------
def read_raw_excel(file_bytes,filename):
    ext=os.path.splitext(filename)[1].lower()
    if ext==".xlsx":
        wb=openpyxl.load_workbook(io.BytesIO(file_bytes),data_only=True,read_only=True)
        ws=wb.worksheets[0]
        raw=[[_cell_text(cell.value) for cell in row] for row in ws.iter_rows()]
    elif ext==".xls":
        wb=xlrd.open_workbook(file_contents=file_bytes); ws=wb.sheet_by_index(0)
        raw=[]
        for r in range(ws.nrows):
            out=[]
            for c in range(ws.ncols):
                cell=ws.cell(r,c)
                val=cell.value
                if cell.ctype==xlrd.XL_CELL_DATE:
                    try: val=xlrd.xldate.xldate_as_datetime(val,wb.datemode)
                    except Exception: pass
                out.append(_cell_text(val))
            raw.append(out)
    else:
        raise ValueError("Format file harus .xls atau .xlsx.")
    if not raw: raise ValueError("File kosong.")
    hrow=0
    for i,row in enumerate(raw[:10]):
        if sum(1 for v in row if str(v).strip())>=3: hrow=i; break
    headers=[str(v).strip() for v in raw[hrow]]
    rows=[[str(v).strip() if v is not None else "" for v in row]
          for row in raw[hrow+1:] if any(str(v).strip() for v in row)]
    return headers,rows

def apply_mapping_and_process(headers,rows,mapping):
    def get(row,field):
        idx=mapping.get(field)
        return str(row[idx]).strip() if idx is not None and idx<len(row) else ""
    cleaned=[]
    for row in rows:
        faktur=get(row,"No. Faktur")
        if not faktur or faktur=="No. Faktur": continue
        qty=parse_num(get(row,"Kuantitas"))
        harga=parse_num(get(row,"Harga Satuan"))
        if qty==0 and harga==0: continue
        raw_tgl=get(row,"Tgl Faktur")
        cleaned.append({"nama":get(row,"Nama Pelanggan"),"keterangan":get(row,"Keterangan Barang"),
                         "faktur":faktur,"tgl":_date_to_ddmmyyyy(raw_tgl),"tgl_raw":raw_tgl,
                         "qty":round2(qty),"harga":round2(harga),
                         "unit":get(row,"Unit 1 Barang"),"detail_opt":"A",
                         "detail_code":"000000","detail_unit":"UM.0021"})
    if not cleaned: raise ValueError("Tidak ada data valid.")
    fmap={}; ctr=1
    for r in cleaned:
        if r["faktur"] not in fmap: fmap[r["faktur"]]=ctr; ctr+=1
        r["no"]=fmap[r["faktur"]]
    for r in cleaned:
        apply_tax_calculation(r,0)
    fc=Counter(r["faktur"] for r in cleaned)
    dup={f for f,n in fc.items() if n>1}
    dup_map={f:i%len(DUP_COLORS) for i,f in enumerate(sorted(dup))}
    return cleaned,dup_map

def calc_totals(processed_data):
    total_qty=sum(r["qty"] for r in processed_data)
    total_diskon=sum(r["diskon"] for r in processed_data)
    total_dpp=sum(r["dpp"] for r in processed_data)
    total_dpp_nl=sum(r["dpp_nl"] for r in processed_data)
    total_ppn=sum(r["ppn"] for r in processed_data)
    return {"qty":total_qty,"diskon":total_diskon,"dpp":total_dpp,"dpp_nl":total_dpp_nl,"ppn":total_ppn}

def apply_dlm_mapping_and_process(headers,rows,mapping):
    def get(row,field):
        idx=mapping.get(field)
        return str(row[idx]).strip() if idx is not None and idx<len(row) else ""
    cleaned=[]
    for row in rows:
        doc_no=get(row,"No. Faktur")
        if not doc_no or doc_no=="No. Faktur": continue
        nilai=parse_num(get(row,"Nilai Faktur"))
        if nilai==0: continue
        doc_date_iso=_date_to_iso(get(row,"Tgl Faktur"))
        m=re.match(r"(\d{4})-(\d{2})-(\d{2})$",doc_date_iso or "")
        month,year=(int(m.group(2)),int(m.group(1))) if m else (0,0)
        cleaned.append({"doc_no":doc_no,"doc_date":doc_date_iso,
                         "trx_type":"04","trx_code":"01","trx_document":"8",
                         "tax_period_month":month,"tax_period_year":year,
                         "seller_name":get(row,"Nama Pemasok"),"seller_tin":"",
                         "nilai_faktur":round2(nilai)})
    if not cleaned: raise ValueError("Tidak ada data valid.")
    for i,r in enumerate(cleaned,1):
        r["no"]=i
        apply_dlm_calculation(r)
    return cleaned

def calc_dlm_totals(processed_data):
    return {"tax_base":sum(r["tax_base"] for r in processed_data),
            "vat":sum(r["vat"] for r in processed_data),
            "stlg":sum(r.get("stlg",0) for r in processed_data)}

def calc_sdl_totals(processed_data):
    return {"tax_base":sum(r["tax_base"] for r in processed_data),
            "other_tax_base":sum(r["other_tax_base"] for r in processed_data),
            "vat":sum(r["vat"] for r in processed_data),
            "stlg":sum(r.get("stlg",0) for r in processed_data)}

def collect_sdl_export_warnings(sdl):
    processed=sdl["processed"]; warns=[]
    npwp=str(sdl.get("company_npwp","")).strip()
    if not npwp: warns.append("NPWP Perusahaan (TIN) belum diisi")
    elif len(npwp)!=16: warns.append("NPWP Perusahaan (TIN) harus 16 digit")
    month=int(sdl.get("tax_period_month") or 0)
    year=int(sdl.get("tax_period_year") or 0)
    if not (1<=month<=12) or not year:
        warns.append("Masa/Tahun Pajak SPT belum diisi")
    n=sum(1 for r in processed if not re.match(r"\d{4}-\d{2}-\d{2}$",str(r.get("doc_date","")).strip()))
    if n: warns.append(f"Tgl Transaksi tidak terbaca: {n} baris")
    if 1<=month<=12 and year:
        n=sum(1 for r in processed
              if (m:=re.match(r"(\d{4})-(\d{2})-\d{2}$",str(r.get("doc_date","")).strip()))
              and (int(m.group(2))!=month or int(m.group(1))!=year))
        if n: warns.append(f"Tgl Transaksi di luar Masa Pajak SPT: {n} baris")
    n=sum(1 for r in processed if not str(r.get("buyer_id_number","")).strip())
    if n: warns.append(f"NPWP/NIK Pembeli kosong: {n} baris")
    n=sum(1 for r in processed if str(r.get("buyer_id_number","")).strip() and len(str(r.get("buyer_id_number","")).strip())!=16)
    if n: warns.append(f"NPWP/NIK Pembeli harus 16 digit: {n} baris")
    n=sum(1 for r in processed if not str(r.get("buyer_name","")).strip())
    if n: warns.append(f"Nama Pembeli kosong: {n} baris")
    n=sum(1 for r in processed if not str(r.get("serial_no","")).strip())
    if n: warns.append(f"SerialNo kosong: {n} baris")
    n=sum(1 for r in processed if str(r.get("trx_code",""))=="NoVAT" and round2(r.get("vat",0))!=0)
    if n: warns.append(f"TrxCode NoVAT tapi nominal PPN bukan 0: {n} baris")
    n=sum(1 for r in processed if str(r.get("trx_code","")).strip() not in SDL_TRX_CODE_LABELS)
    if n: warns.append(f"TrxCode di luar referensi: {n} baris")
    n=sum(1 for r in processed if str(r.get("buyer_id_opt","")).strip() not in SDL_BUYER_ID_LABELS)
    if n: warns.append(f"BuyerIdOpt di luar referensi: {n} baris")
    n=sum(1 for r in processed if str(r.get("good_service_opt","")).strip() not in SDL_GOOD_SERVICE_LABELS)
    if n: warns.append(f"GoodServiceOpt di luar referensi: {n} baris")
    return warns

def collect_dlm_export_warnings(dlm):
    processed=dlm["processed"]; warns=[]
    npwp=str(dlm.get("company_npwp","")).strip()
    if not npwp: warns.append("NPWP Perusahaan (TIN) belum diisi")
    elif len(npwp)!=16: warns.append("NPWP Perusahaan (TIN) harus 16 digit")
    n=sum(1 for r in processed if not str(r.get("seller_tin","")).strip())
    if n: warns.append(f"NPWP Pemasok kosong: {n} baris")
    n=sum(1 for r in processed if str(r.get("seller_tin","")).strip() and len(str(r.get("seller_tin","")).strip())!=16)
    if n: warns.append(f"NPWP Pemasok harus 16 digit: {n} baris")
    n=sum(1 for r in processed if not r.get("tax_period_month") or not r.get("tax_period_year"))
    if n: warns.append(f"Masa/Tahun Pajak tidak terbaca: {n} baris")
    n=sum(1 for r in processed if str(r.get("trx_type","")).strip() not in DLM_TRX_TYPE_LABELS)
    if n: warns.append(f"TrxType di luar referensi: {n} baris")
    n=sum(1 for r in processed if str(r.get("trx_code","")).strip() not in DLM_TRX_DETAIL_LABELS)
    if n: warns.append(f"TrxCode di luar referensi: {n} baris")
    n=sum(1 for r in processed if str(r.get("trx_document","")).strip() not in DLM_TRX_DOC_LABELS)
    if n: warns.append(f"TrxDocument di luar referensi: {n} baris")
    return warns

def build_faktur_rows(processed_data):
    seen=set(); rows=[]; baris=1
    for r in processed_data:
        if r["faktur"] in seen: continue
        seen.add(r["faktur"])
        rows.append({"baris":str(baris),"tgl_faktur":r["tgl"],"jenis_faktur":"Normal",
                      "tgl_raw":r.get("tgl_raw",""),
                      "kode_transaksi":"04","ket_tambahan":"","dok_pendukung":"",
                      "referensi":r["faktur"],"cap_fasilitas":"","id_tku_penjual":"",
                      "npwp_pembeli":"","jenis_id":"TIN","negara":"IDN","no_dok_pembeli":"",
                      "nama_pembeli":"","alamat_pembeli":"","email_pembeli":"",
                      "id_tku_pembeli":"","diskon_rate":0})
        baris+=1
    return rows

def iter_detail_rows_by_faktur(processed_data,faktur_rows):
    ref_to_baris={r["referensi"]:r["baris"] for r in faktur_rows}
    ref_order={r["referensi"]:i for i,r in enumerate(faktur_rows)}
    fallback=len(ref_order)
    indexed=list(enumerate(processed_data))
    indexed.sort(key=lambda item:(ref_order.get(item[1]["faktur"],fallback+item[0]),item[0]))
    for original_idx,r in indexed:
        yield original_idx,r,ref_to_baris.get(r["faktur"],str(original_idx+1))

def build_detail_rows(processed_data,faktur_rows):
    rows=[]
    for original_idx,r,bn in iter_detail_rows_by_faktur(processed_data,faktur_rows):
        rows.append({
            "original_idx":original_idx,"baris":bn,
            "detail_opt":r.get("detail_opt","A"),
            "detail_opt_label":_detail_opt_label(r.get("detail_opt","A")),
            "detail_code":r.get("detail_code","000000"),
            "keterangan":r["keterangan"],
            "detail_unit":r.get("detail_unit","UM.0021"),
            "detail_unit_label":_detail_unit_label(r.get("detail_unit","UM.0021")),
            "harga":round2(r["harga"]),"qty":round2(r["qty"]),
            "diskon":round2(r.get("diskon",0)),"dpp":round2(r["dpp"]),
            "dpp_nl":round2(r["dpp_nl"]),"ppn":round2(r["ppn"]),
        })
    return rows

def collect_export_warnings(faktur_rows,processed):
    warns=[]
    def count(rows,pred):
        return sum(1 for r in rows if pred(r))

    warns.extend(collect_faktur_date_diagnostics(faktur_rows)["warnings"])
    n=count(faktur_rows,lambda r:not _is_ddmmyyyy(r.get("tgl_faktur","")))
    if n: warns.append(f"Tanggal Faktur bukan format DD/MM/YYYY: {n} baris")
    n=count(faktur_rows,lambda r:len(str(r.get("id_tku_penjual","")).strip())!=22)
    if n: warns.append(f"ID TKU Penjual harus 22 digit: {n} baris")
    n=count(faktur_rows,lambda r:not str(r.get("npwp_pembeli","")).strip())
    if n: warns.append(f"NPWP/NIK Pembeli kosong: {n} baris")
    n=count(faktur_rows,lambda r:str(r.get("jenis_id","TIN")).strip()=="TIN" and len(str(r.get("npwp_pembeli","")).strip())!=16)
    if n: warns.append(f"NPWP/NIK Pembeli TIN harus 16 digit: {n} baris")
    n=count(faktur_rows,lambda r:not str(r.get("nama_pembeli","")).strip())
    if n: warns.append(f"Nama Pembeli kosong: {n} baris")
    n=count(faktur_rows,lambda r:not str(r.get("alamat_pembeli","")).strip())
    if n: warns.append(f"Alamat Pembeli kosong: {n} baris")
    n=count(faktur_rows,lambda r:str(r.get("jenis_id","TIN")).strip()=="TIN" and len(str(r.get("id_tku_pembeli","")).strip())!=22)
    if n: warns.append(f"ID TKU Pembeli TIN harus 22 digit: {n} baris")
    n=count(faktur_rows,lambda r:str(r.get("jenis_id","TIN")).strip()!="TIN" and not str(r.get("no_dok_pembeli","")).strip())
    if n: warns.append(f"Nomor Dokumen Pembeli wajib untuk non-TIN: {n} baris")

    n=count(processed,lambda r:r.get("detail_opt","A") not in DETAIL_OPT_LABELS)
    if n: warns.append(f"Barang/Jasa detail di luar referensi A/B: {n} baris")
    n=count(processed,lambda r:r.get("detail_unit","UM.0021") not in DETAIL_UNIT_LABELS)
    if n: warns.append(f"Satuan detail di luar referensi: {n} baris")
    n=count(processed,lambda r:not str(r.get("keterangan","")).strip())
    if n: warns.append(f"Nama Barang/Jasa kosong: {n} baris")
    n=count(processed,lambda r:round2(r.get("harga",0))<=0)
    if n: warns.append(f"Harga Satuan harus lebih dari 0: {n} baris")
    n=count(processed,lambda r:round2(r.get("qty",0))<=0)
    if n: warns.append(f"Jumlah Barang/Jasa harus lebih dari 0: {n} baris")
    return warns

def collect_export_blockers(faktur_rows):
    blockers=[]
    def count(rows,pred):
        return sum(1 for r in rows if pred(r))

    blockers.extend(collect_faktur_date_diagnostics(faktur_rows)["blockers"])
    n=count(faktur_rows,lambda r:len(str(r.get("id_tku_penjual","")).strip())!=22)
    if n: blockers.append(f"ID TKU Penjual wajib diisi 22 digit: {n} baris")
    n=count(faktur_rows,lambda r:not str(r.get("npwp_pembeli","")).strip())
    if n: blockers.append(f"NPWP/NIK Pembeli wajib diisi: {n} baris")
    n=count(faktur_rows,lambda r:str(r.get("jenis_id","TIN")).strip()=="TIN" and len(str(r.get("npwp_pembeli","")).strip())!=16)
    if n: blockers.append(f"NPWP/NIK Pembeli TIN wajib 16 digit: {n} baris")
    n=count(faktur_rows,lambda r:not str(r.get("nama_pembeli","")).strip())
    if n: blockers.append(f"Nama Pembeli wajib diisi: {n} baris")
    n=count(faktur_rows,lambda r:not str(r.get("alamat_pembeli","")).strip())
    if n: blockers.append(f"Alamat Pembeli wajib diisi: {n} baris")
    n=count(faktur_rows,lambda r:str(r.get("jenis_id","TIN")).strip()=="TIN" and len(str(r.get("id_tku_pembeli","")).strip())!=22)
    if n: blockers.append(f"ID TKU Pembeli TIN wajib 22 digit: {n} baris")
    n=count(faktur_rows,lambda r:str(r.get("jenis_id","TIN")).strip()!="TIN" and not str(r.get("no_dok_pembeli","")).strip())
    if n: blockers.append(f"Nomor Dokumen Pembeli wajib untuk non-TIN: {n} baris")
    return blockers

def apply_buyer_discounts_to_active_data(state,db_pembeli,buyer_name=None):
    processed=state["processed"]; faktur_rows=state["faktur_rows"]
    if not processed: return 0
    targets=[]
    for name,d in db_pembeli.items():
        if buyer_name and norm_party_name(name)!=norm_party_name(buyer_name): continue
        pct=parse_num(d.get("diskon_pct",""))
        names={norm_party_name(name),norm_party_name(d.get("nama",""))}
        names={n for n in names if n}
        targets.append((names,pct))
    if not targets: return 0

    faktur_to_buyer={}
    for fr in faktur_rows:
        buyer=norm_party_name(fr.get("nama_pembeli",""))
        if buyer: faktur_to_buyer[fr["referensi"]]=buyer

    changed=0
    for r in processed:
        src_name=norm_party_name(r.get("nama",""))
        buyer_name_for_ref=faktur_to_buyer.get(r["faktur"],"")
        for names,pct in targets:
            if src_name in names or buyer_name_for_ref in names:
                apply_tax_calculation(r,pct)
                changed+=1
                break

    for fr in faktur_rows:
        buyer=norm_party_name(fr.get("nama_pembeli",""))
        if not buyer:
            src=next((norm_party_name(r.get("nama","")) for r in processed if r["faktur"]==fr["referensi"]),"")
            buyer=src
        for names,pct in targets:
            if buyer in names:
                fr["diskon_rate"]=pct
                break

    if changed:
        state["totals"]=calc_totals(processed)
    return changed

def recalc_detail_for_refs(state,refs,diskon_pct):
    refset=set(refs)
    for r in state["processed"]:
        if r["faktur"] in refset:
            apply_tax_calculation(r,diskon_pct)
    state["totals"]=calc_totals(state["processed"])

def apply_invoice_date_to_refs(state,refs,tgl):
    refset=set(refs)
    for r in state["processed"]:
        if r["faktur"] in refset:
            r["tgl"]=tgl


# ---- Export XLSX ----------------------------------------------------------------
def _brd():
    t=Side(border_style="thin",color="000000")
    return Border(left=t,right=t,top=t,bottom=t)

def _al(h="center",v="center"):
    return Alignment(horizontal=h,vertical=v,wrap_text=False)

def export_tpk_sheet(ws,data,dup_color_map):
    brd=_brd(); ca=_al(); la=_al("left"); ra=_al("right")
    HDRS=["No","Nama Pelanggan","Keterangan Barang","No. Faktur","Tgl Faktur",
          "Kuantitas","Harga satuan","Unit 1 Barang","Harga Jual","Diskon",
          "DPP","DPP Nilai Lain","Tarif PPN","PPN","Tarif PPnBM","PPnBM"]
    hf=Font(name="Arial",size=9,bold=True,color="000080")
    for ci,h in enumerate(HDRS,1):
        c=ws.cell(row=1,column=ci,value=h); c.font=hf; c.alignment=ca; c.border=brd
    ws.row_dimensions[1].height=17
    a8=Font(name="Arial",size=8); c11=Font(name="Calibri",size=11)
    for i,r in enumerate(data):
        er=i+2; dup=r["faktur"] in dup_color_map; ci=dup_color_map.get(r["faktur"])
        rf=PatternFill("solid",fgColor=DUP_COLORS[ci][1].lstrip("#")) if dup else None
        def wc(col,val,font=None,align=None,nf=None,fill=None,_er=er,_rf=rf):
            cell=ws.cell(row=_er,column=col,value=val)
            cell.font=font or a8; cell.alignment=align or ca; cell.border=brd
            if nf: cell.number_format=nf
            if fill: cell.fill=fill
            elif _rf: cell.fill=_rf
        wc(1,str(r["no"]),font=c11,nf="@"); wc(2,r["nama"],align=la); wc(3,r["keterangan"],align=ca)
        ff=Font(name="Arial",size=8,color=DUP_COLORS[ci][0].lstrip("#") if dup else "000000")
        c4=ws.cell(row=er,column=4,value=r["faktur"]); c4.font=ff; c4.alignment=ca; c4.border=brd
        if dup: c4.fill=PatternFill("solid",fgColor=DUP_COLORS[ci][1].lstrip("#"))
        wc(5,r["tgl"]); wc(6,round2(r["qty"]),align=ra,nf="#,##0.##")
        wc(7,round2(r["harga"]),align=ra,nf="#,##0.##"); wc(8,r["unit"])
        for col,val in [(9,r["harga_jual"]),(10,r["diskon"]),(11,r["dpp"]),(12,r["dpp_nl"]),(14,r["ppn"])]:
            c=ws.cell(row=er,column=col,value=round2(val)); c.font=c11; c.alignment=ra; c.border=brd
            c.number_format="#,##0.00"
            if rf: c.fill=rf
        c13=ws.cell(row=er,column=13,value=12); c13.font=c11; c13.alignment=ca; c13.border=brd
        if rf: c13.fill=rf
        for col in (15,16):
            c=ws.cell(row=er,column=col,value=None); c.font=c11; c.border=brd
            if rf: c.fill=rf
        ws.row_dimensions[er].height=16
    autofit_columns(ws); ws.freeze_panes="A2"

def export_dlm_sheet(ws,data,company_npwp=""):
    brd=_brd(); ca=_al(); la=_al("left"); ra=_al("right")
    ws.cell(row=1,column=1,value="NPWP")
    c=ws.cell(row=1,column=2,value=company_npwp); c.number_format="@"
    HDRS=["SpecialDocNo","SpecialDocDate","TrxType","TrxCode","TrxDocument",
          "TaxPeriodMonth","TaxPeriodYear","TotalTaxBase","TotalVAT","TotalSTLG",
          "SellerTin","SellerName"]
    hf=Font(name="Calibri",size=11,bold=True)
    for ci,h in enumerate(HDRS,1):
        c=ws.cell(row=3,column=ci,value=h); c.font=hf; c.alignment=ca; c.border=brd
    c11=Font(name="Calibri",size=11)
    for i,r in enumerate(data):
        er=i+4
        c=ws.cell(row=er,column=1,value=r["doc_no"]); c.font=c11; c.alignment=ca; c.border=brd; c.number_format="@"
        iso=r.get("doc_date","")
        m=re.match(r"(\d{4})-(\d{2})-(\d{2})$",iso or "")
        dval=datetime(int(m.group(1)),int(m.group(2)),int(m.group(3))) if m else iso
        c=ws.cell(row=er,column=2,value=dval); c.font=c11; c.alignment=ca; c.border=brd
        if m: c.number_format="dd/mm/yyyy"
        c=ws.cell(row=er,column=3,value=r["trx_type"]); c.font=c11; c.alignment=ca; c.border=brd; c.number_format="@"
        c=ws.cell(row=er,column=4,value=r["trx_code"]); c.font=c11; c.alignment=ca; c.border=brd; c.number_format="@"
        td=r["trx_document"]
        c=ws.cell(row=er,column=5,value=int(td) if str(td).isdigit() else td); c.font=c11; c.alignment=ca; c.border=brd
        c=ws.cell(row=er,column=6,value=r["tax_period_month"]); c.font=c11; c.alignment=ca; c.border=brd
        c=ws.cell(row=er,column=7,value=r["tax_period_year"]); c.font=c11; c.alignment=ca; c.border=brd
        c=ws.cell(row=er,column=8,value=round2(r["tax_base"])); c.font=c11; c.alignment=ra; c.border=brd; c.number_format="#,##0"
        c=ws.cell(row=er,column=9,value=round2(r["vat"])); c.font=c11; c.alignment=ra; c.border=brd; c.number_format="#,##0"
        c=ws.cell(row=er,column=10,value=round2(r.get("stlg",0))); c.font=c11; c.alignment=ra; c.border=brd; c.number_format="#,##0"
        c=ws.cell(row=er,column=11,value=r.get("seller_tin","")); c.font=c11; c.alignment=ca; c.border=brd; c.number_format="@"
        c=ws.cell(row=er,column=12,value=r.get("seller_name","")); c.font=c11; c.alignment=la; c.border=brd
        ws.row_dimensions[er].height=16
    if data:
        last_row=len(data)+3
        table=Table(displayName="Table2",ref=f"A3:L{last_row}")
        table.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showFirstColumn=False,
                                            showLastColumn=False,showRowStripes=True,showColumnStripes=False)
        ws.add_table(table)
        refs=[
            ("C",f"='Trx TYpe'!$A$2:$A${len(DLM_TRX_TYPE_LABELS)+1}"),
            ("D",f"='Trx Detail'!$A$2:$A${len(DLM_TRX_DETAIL_LABELS)+1}"),
            ("E",f"='Trx Doc'!$A$2:$A${len(DLM_TRX_DOC_LABELS)+1}"),
        ]
        for col,formula in refs:
            dv=DataValidation(type="list",formula1=formula,allow_blank=False)
            ws.add_data_validation(dv)
            dv.add(f"{col}4:{col}{last_row}")
    autofit_columns(ws); ws.freeze_panes="A4"

def export_dlm_reference_sheet(ws,title,labels):
    brd=_brd(); ca=_al(); la=_al("left")
    ws.title=title
    hf=Font(name="Calibri",size=11,bold=True)
    c11=Font(name="Calibri",size=11)
    for ci,h in enumerate(("Kode","Deskripsi"),1):
        c=ws.cell(row=1,column=ci,value=h); c.font=hf; c.alignment=ca; c.border=brd
    for ri,(code,desc) in enumerate(labels.items(),2):
        c=ws.cell(row=ri,column=1,value=code); c.font=c11; c.alignment=ca; c.border=brd; c.number_format="@"
        c=ws.cell(row=ri,column=2,value=desc); c.font=c11; c.alignment=la; c.border=brd
    autofit_columns(ws); ws.freeze_panes="A2"

def export_faktur_sheet(ws,faktur_rows,npwp_penjual=""):
    brd=_brd(); ca=_al(); la=_al("left")
    ws.cell(row=1,column=1,value="NPWP Penjual")
    c=ws.cell(row=1,column=3,value=npwp_penjual); c.number_format="@"
    HDRS=["Baris","Tanggal Faktur","Jenis Faktur","Kode Transaksi","Keterangan Tambahan",
          "Dokumen Pendukung","Referensi","Cap Fasilitas","ID TKU Penjual","NPWP/NIK Pembeli",
          "Jenis ID Pembeli","Negara Pembeli","Nomor Dokumen Pembeli","Nama Pembeli",
          "Alamat Pembeli","Email Pembeli","ID TKU Pembeli"]
    hf=Font(name="Calibri",size=11,bold=True)
    for ci,h in enumerate(HDRS,1):
        c=ws.cell(row=3,column=ci,value=h); c.font=hf; c.alignment=ca; c.border=brd
    c11=Font(name="Calibri",size=11)
    for i,r in enumerate(faktur_rows):
        er=i+4
        def wt(col,val,_er=er):
            c=ws.cell(row=_er,column=col,value=str(val) if val else "")
            c.font=c11; c.alignment=ca; c.border=brd; c.number_format="@"
        def wl(col,val,_er=er):
            c=ws.cell(row=_er,column=col,value=str(val) if val else "")
            c.font=c11; c.alignment=la; c.border=brd
        wt(1,r["baris"]); wt(2,r["tgl_faktur"]); wt(3,r["jenis_faktur"]); wt(4,r["kode_transaksi"])
        wt(5,r["ket_tambahan"]); wt(6,r["dok_pendukung"]); wt(7,r["referensi"]); wt(8,r["cap_fasilitas"])
        wt(9,r["id_tku_penjual"]); wt(10,r["npwp_pembeli"]); wt(11,r["jenis_id"]); wt(12,r["negara"])
        wt(13,r["no_dok_pembeli"]); wl(14,r["nama_pembeli"]); wl(15,r["alamat_pembeli"])
        wt(16,r["email_pembeli"]); wt(17,r["id_tku_pembeli"])
        ws.row_dimensions[er].height=16
    end_row=len(faktur_rows)+4
    ce=ws.cell(row=end_row,column=1,value="END")
    ce.font=Font(name="Calibri",size=11,bold=True)
    ce.alignment=ca; ce.border=brd; ce.number_format="@"
    for col in range(2,18):
        c=ws.cell(row=end_row,column=col,value=None)
        c.font=c11; c.alignment=ca; c.border=brd
    ws.row_dimensions[end_row].height=16
    autofit_columns(ws); ws.freeze_panes="A4"

def export_detail_faktur_sheet(ws,processed_data,faktur_rows):
    brd=_brd(); ca=_al(); la=_al("left"); ra=_al("right")
    HDRS=["Baris","Barang/Jasa","Kode Barang Jasa","Nama Barang/Jasa","Nama Satuan Ukur",
          "Harga Satuan","Jumlah Barang Jasa","Total Diskon","DPP","DPP Nilai Lain",
          "Tarif PPN","PPN","Tarif PPnBM","PPnBM"]
    hf=Font(name="Calibri",size=11,bold=True)
    for ci,h in enumerate(HDRS,1):
        c=ws.cell(row=1,column=ci,value=h); c.font=hf; c.alignment=ca; c.border=brd
    ws.row_dimensions[1].height=16
    ordered_details=list(iter_detail_rows_by_faktur(processed_data,faktur_rows))
    c11=Font(name="Calibri",size=11)
    for i,(_original_idx,r,bn) in enumerate(ordered_details):
        er=i+2
        def wt(col,val,_er=er):
            c=ws.cell(row=_er,column=col,value=str(val) if val is not None else "")
            c.font=c11; c.alignment=ca; c.border=brd; c.number_format="@"
        def wn(col,val,_er=er):
            c=ws.cell(row=_er,column=col,value=val); c.font=c11; c.alignment=ra; c.border=brd
            c.number_format="#,##0.00"
        def wl(col,val,_er=er):
            c=ws.cell(row=_er,column=col,value=str(val) if val else "")
            c.font=c11; c.alignment=la; c.border=brd
        wt(1,bn); wt(2,r.get("detail_opt","A")); wt(3,r.get("detail_code","000000"))
        wl(4,r["keterangan"]); wt(5,r.get("detail_unit","UM.0021"))
        wn(6,round2(r["harga"])); wn(7,round2(r["qty"]))
        wn(8,round2(r.get("diskon",0))); wn(9,round2(r["dpp"]))
        wn(10,round2(r["dpp_nl"])); wn(11,12); wn(12,round2(r["ppn"]))
        wn(13,0); wn(14,0)
        ws.row_dimensions[er].height=16
    end_row=len(ordered_details)+2
    ce=ws.cell(row=end_row,column=1,value="END")
    ce.font=Font(name="Calibri",size=11,bold=True)
    ce.alignment=ca; ce.border=brd; ce.number_format="@"
    for col in range(2,15):
        c=ws.cell(row=end_row,column=col,value=None); c.font=c11; c.border=brd
    ws.row_dimensions[end_row].height=16
    autofit_columns(ws); ws.freeze_panes="A2"


# ---- Generate XML -----------------------------------------------------------------
def generate_coretax_xml(processed_data,faktur_rows):
    ref_map={r["referensi"]:r for r in faktur_rows}
    goods_map=defaultdict(list)
    ordered_details=list(iter_detail_rows_by_faktur(processed_data,faktur_rows))
    for _original_idx,r,_baris in ordered_details:
        goods_map[r["faktur"]].append(r)
    tin_seller=""
    if faktur_rows:
        id_tku=faktur_rows[0].get("id_tku_penjual","")
        tin_seller=id_tku[:16] if len(id_tku)>=16 else id_tku
    root=Element("TaxInvoiceBulk")
    root.set("xmlns:xsd","http://www.w3.org/2001/XMLSchema")
    root.set("xmlns:xsi","http://www.w3.org/2001/XMLSchema-instance")
    SubElement(root,"TIN").text=tin_seller
    list_inv=SubElement(root,"ListOfTaxInvoice")
    seen=set(); ordered_refs=[]
    for fr in faktur_rows:
        ref=fr.get("referensi","")
        if ref and ref in goods_map and ref not in seen:
            seen.add(ref); ordered_refs.append(ref)
    for _original_idx,r,_baris in ordered_details:
        if r["faktur"] not in seen:
            seen.add(r["faktur"]); ordered_refs.append(r["faktur"])
    for ref in ordered_refs:
        fr=ref_map.get(ref,{}); inv=SubElement(list_inv,"TaxInvoice")
        SubElement(inv,"TaxInvoiceDate").text=_date_to_iso(fr.get("tgl_faktur",""))
        SubElement(inv,"TaxInvoiceOpt").text=fr.get("jenis_faktur","Normal")
        SubElement(inv,"TrxCode").text=fr.get("kode_transaksi","04")
        SubElement(inv,"AddInfo").text=fr.get("ket_tambahan","")
        SubElement(inv,"CustomDoc").text=fr.get("dok_pendukung","")
        SubElement(inv,"RefDesc").text=ref
        SubElement(inv,"FacilityStamp").text=fr.get("cap_fasilitas","")
        SubElement(inv,"SellerIDTKU").text=fr.get("id_tku_penjual","")
        SubElement(inv,"BuyerTin").text=fr.get("npwp_pembeli","")
        SubElement(inv,"BuyerDocument").text=fr.get("jenis_id","TIN")
        SubElement(inv,"BuyerCountry").text=fr.get("negara","IDN")
        SubElement(inv,"BuyerDocumentNumber").text=fr.get("no_dok_pembeli","")
        SubElement(inv,"BuyerName").text=fr.get("nama_pembeli","")
        SubElement(inv,"BuyerAdress").text=fr.get("alamat_pembeli","")
        SubElement(inv,"BuyerEmail").text=fr.get("email_pembeli","")
        SubElement(inv,"BuyerIDTKU").text=fr.get("id_tku_pembeli","")
        list_gs=SubElement(inv,"ListOfGoodService")
        for r in goods_map[ref]:
            gs=SubElement(list_gs,"GoodService")
            SubElement(gs,"Opt").text=r.get("detail_opt","A")
            SubElement(gs,"Code").text=r.get("detail_code","000000")
            SubElement(gs,"Name").text=r["keterangan"]
            SubElement(gs,"Unit").text=r.get("detail_unit","UM.0021")
            SubElement(gs,"Price").text=xml_num(r["harga"]); SubElement(gs,"Qty").text=xml_num(r["qty"])
            SubElement(gs,"TotalDiscount").text=xml_num(r.get("diskon",0))
            SubElement(gs,"TaxBase").text=xml_num(r["dpp"])
            SubElement(gs,"OtherTaxBase").text=xml_num(r["dpp_nl"])
            SubElement(gs,"VATRate").text="12"; SubElement(gs,"VAT").text=xml_num(r["ppn"])
            SubElement(gs,"STLGRate").text="0"; SubElement(gs,"STLG").text="0"
    raw=tostring(root,encoding="unicode")
    pretty=minidom.parseString(raw).toprettyxml(indent="  ",encoding="utf-8").decode("utf-8")
    return pretty

def generate_dlm_xml(processed_data,company_npwp=""):
    root=Element("InputSpecialDocBulk")
    SubElement(root,"TIN").text=company_npwp
    list_doc=SubElement(root,"ListOfSpecialDoc")
    for r in processed_data:
        sd=SubElement(list_doc,"SpecialDoc")
        SubElement(sd,"SpecialDocNo").text=r["doc_no"]
        SubElement(sd,"SpecialDocDate").text=r.get("doc_date","")
        SubElement(sd,"TrxType").text=r["trx_type"]
        SubElement(sd,"TrxCode").text=r["trx_code"]
        SubElement(sd,"TrxDocument").text=str(r["trx_document"])
        SubElement(sd,"TaxPeriodMonth").text=str(r["tax_period_month"])
        SubElement(sd,"TaxPeriodYear").text=str(r["tax_period_year"])
        SubElement(sd,"TotalTaxBase").text=xml_num_min(r["tax_base"])
        SubElement(sd,"TotalVAT").text=xml_num_min(r["vat"])
        SubElement(sd,"TotalSTLG").text=xml_num_min(r.get("stlg",0))
        SubElement(sd,"SellerTin").text=r.get("seller_tin","")
        SubElement(sd,"SellerName").text=r.get("seller_name","")
    raw=tostring(root,encoding="unicode")
    body=minidom.parseString(raw).toprettyxml(indent="\t")
    body=body.split("\n",1)[1]
    return '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'+body

def generate_sdl_xml(processed_data,company_npwp="",month=0,year=0):
    root=Element("RetailInvoiceBulk")
    SubElement(root,"TIN").text=company_npwp
    SubElement(root,"TaxPeriodMonth").text=f"{int(month or 0):02d}"
    SubElement(root,"TaxPeriodYear").text=str(int(year or 0))
    list_inv=SubElement(root,"ListOfRetailInvoice")
    for r in processed_data:
        ri=SubElement(list_inv,"RetailInvoice")
        SubElement(ri,"TrxCode").text=r["trx_code"]
        SubElement(ri,"BuyerName").text=r.get("buyer_name","")
        SubElement(ri,"BuyerIdOpt").text=r.get("buyer_id_opt","NPWP")
        SubElement(ri,"BuyerIdNumber").text=r.get("buyer_id_number","")
        SubElement(ri,"GoodServiceOpt").text=r.get("good_service_opt","A")
        SubElement(ri,"SerialNo").text=r["serial_no"]
        SubElement(ri,"TransactionDate").text=r.get("doc_date","")
        SubElement(ri,"TaxBaseSellingPrice").text=xml_num_min(r["tax_base"])
        SubElement(ri,"OtherTaxBaseSellingPrice").text=xml_num_min(r["other_tax_base"])
        SubElement(ri,"VAT").text=xml_num_min(r["vat"])
        SubElement(ri,"STLG").text=xml_num_min(r.get("stlg",0))
        SubElement(ri,"Info").text=str(r.get("info","") or "0")
    raw=tostring(root,encoding="unicode")
    body=minidom.parseString(raw).toprettyxml(indent="\t")
    body=body.split("\n",1)[1]
    return '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'+body


# ==============================================================================
#  FLASK APP
# ==============================================================================
def _get_secret_key():
    configured = os.environ.get("FLASK_SECRET_KEY") or os.environ.get("SECRET_KEY")
    if configured:
        return configured
    if os.path.exists(SECRET_KEY_FILE):
        try: return open(SECRET_KEY_FILE,"r").read().strip()
        except Exception: pass
    key=uuid.uuid4().hex
    try:
        with open(SECRET_KEY_FILE,"w") as f: f.write(key)
    except OSError:
        pass
    return key

app=Flask(__name__)
app.secret_key=_get_secret_key()
app.json.ensure_ascii=False

SESSIONS={}
TRIAL_INVOICE_LIMIT = int(os.environ.get("TRIAL_INVOICE_LIMIT", "10"))
PRICING_URL = "/auth?plan=professional"

def trial_info(used=0):
    return {
        "plan": "Trial",
        "limit": TRIAL_INVOICE_LIMIT,
        "used": int(used or 0),
        "remaining": max(TRIAL_INVOICE_LIMIT - int(used or 0), 0),
        "pricing_url": PRICING_URL,
        "locked_features": ["doc_lain_masukan", "spt_dokumen_lain"],
    }

def trial_upgrade_response(attempted, module_name="invoice"):
    return jsonify({
        "error": f"Paket Trial hanya bisa memproses maksimal {TRIAL_INVOICE_LIMIT} invoice. Upgrade package untuk memproses {attempted} {module_name}.",
        "upgrade_required": True,
        "plan": "Trial",
        "limit": TRIAL_INVOICE_LIMIT,
        "attempted": int(attempted or 0),
        "pricing_url": PRICING_URL,
        "packages": ["Starter", "Professional", "Enterprise"],
    }), 402

def trial_feature_locked_response(feature_name):
    return jsonify({
        "error": f"Fitur {feature_name} terkunci untuk paket Trial. Login dan pilih package Starter, Professional, atau Enterprise untuk membuka fitur ini.",
        "upgrade_required": True,
        "feature_locked": True,
        "plan": "Trial",
        "limit": TRIAL_INVOICE_LIMIT,
        "attempted": 0,
        "pricing_url": PRICING_URL,
        "packages": ["Starter", "Professional", "Enterprise"],
    }), 402

def default_dlm_state():
    return {
        "raw_headers":[], "raw_rows":[], "mapping":{},
        "src_filename":"", "processed":[], "totals":{},
        "company_npwp":"",
    }

def default_sdl_state():
    return {
        "processed":[], "totals":{},
        "company_npwp":"", "tax_period_month":0, "tax_period_year":0,
    }

def default_state():
    return {
        "raw_headers":[], "raw_rows":[], "mapping":{},
        "src_filename":"", "processed":[], "dup_map":{},
        "faktur_rows":[], "totals":{},
        "dlm": default_dlm_state(),
        "sdl": default_sdl_state(),
    }

def get_state():
    sid=session.get("sid")
    if not sid or sid not in SESSIONS:
        sid=uuid.uuid4().hex
        session["sid"]=sid
        SESSIONS[sid]=default_state()
    st=SESSIONS[sid]
    st.setdefault("sdl",default_sdl_state())
    return st

@app.before_request
def enforce_trial_feature_locks():
    path=request.path or ""
    if path.startswith("/api/dlm/") or path == "/api/dlm":
        return trial_feature_locked_response("Doc Lain Masukan")
    if path.startswith("/api/sdl/") or path == "/api/sdl":
        return trial_feature_locked_response("SPT Dokumen Lain")
    return None

def public_dlm_state(dlm):
    return {
        "raw_headers": dlm["raw_headers"],
        "raw_preview": dlm["raw_rows"][:60],
        "raw_row_count": len(dlm["raw_rows"]),
        "mapping": dlm["mapping"],
        "src_filename": dlm["src_filename"],
        "processed": dlm["processed"],
        "totals": dlm["totals"],
        "company_npwp": dlm.get("company_npwp",""),
    }

def public_sdl_state(sdl):
    return {
        "processed": sdl["processed"],
        "totals": sdl["totals"],
        "company_npwp": sdl.get("company_npwp",""),
        "tax_period_month": sdl.get("tax_period_month",0),
        "tax_period_year": sdl.get("tax_period_year",0),
    }

def public_state(st):
    trial_used=max(len(st.get("faktur_rows",[])),
                   len(st.get("dlm",{}).get("processed",[])),
                   len(st.get("sdl",{}).get("processed",[])))
    return {
        "raw_headers": st["raw_headers"],
        "raw_preview": st["raw_rows"][:60],
        "raw_row_count": len(st["raw_rows"]),
        "mapping": st["mapping"],
        "src_filename": st["src_filename"],
        "processed": st["processed"],
        "dup_map": st["dup_map"],
        "faktur_rows": st["faktur_rows"],
        "date_diagnostics": collect_faktur_date_diagnostics(st["faktur_rows"]),
        "detail_rows": build_detail_rows(st["processed"],st["faktur_rows"]),
        "totals": st["totals"],
        "dlm": public_dlm_state(st["dlm"]),
        "sdl": public_sdl_state(st["sdl"]),
        "subscription": trial_info(trial_used),
    }

def _safe_stem(filename):
    stem=os.path.splitext(os.path.basename(filename or "export"))[0]
    stem=re.sub(r"[^A-Za-z0-9 _.\-]+","_",stem).strip() or "export"
    return stem

def _record_export_history(st):
    history=_jload(EXPORT_HISTORY_FILE)
    now=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for fr in st["faktur_rows"]:
        ref=fr.get("referensi","")
        if not ref: continue
        history[ref]={
            "no_faktur":ref,
            "tgl_faktur":fr.get("tgl_faktur",""),
            "nama_pembeli":fr.get("nama_pembeli",""),
            "source_file":st["src_filename"],
            "exported_at":now,
        }
    _jsave(EXPORT_HISTORY_FILE,history)


# ---- Pages -------------------------------------------------------------------
@app.route("/")
def index():
    return render_template("landing.html")


@app.route("/app")
def app_page():
    return render_template("index.html")


# ---- Bootstrap -----------------------------------------------------------------
@app.route("/api/bootstrap")
def bootstrap():
    st=get_state()
    settings=_jload(SETTINGS_FILE)
    theme=settings.get("theme","Dark Executive")
    if theme not in THEMES: theme="Dark Executive"
    return jsonify({
        "themes": THEMES,
        "theme_names": list(THEMES.keys()),
        "theme": theme,
        "tpk_fields": TPK_FIELDS,
        "field_icons": FIELD_ICONS,
        "out_cols": OUT_COLS,
        "faktur_cols": FAKTUR_COLS,
        "det_cols": DET_COLS,
        "dup_colors": DUP_COLORS,
        "detail_opt_choices": DETAIL_OPT_CHOICES,
        "detail_unit_choices": DETAIL_UNIT_CHOICES,
        "detail_opt_labels": DETAIL_OPT_LABELS,
        "detail_unit_labels": DETAIL_UNIT_LABELS,
        "templates": _jload(TEMPLATES_FILE),
        "db_penjual": _jload(DB_PENJUAL_FILE),
        "db_pembeli": _jload(DB_PEMBELI_FILE),
        "dlm_fields": DLM_FIELDS,
        "dlm_field_icons": DLM_FIELD_ICONS,
        "dlm_out_cols": DLM_OUT_COLS,
        "dlm_trx_type_choices": DLM_TRX_TYPE_CHOICES,
        "dlm_trx_detail_choices": DLM_TRX_DETAIL_CHOICES,
        "dlm_trx_doc_choices": DLM_TRX_DOC_CHOICES,
        "dlm_trx_type_labels": DLM_TRX_TYPE_LABELS,
        "dlm_trx_detail_labels": DLM_TRX_DETAIL_LABELS,
        "dlm_trx_doc_labels": DLM_TRX_DOC_LABELS,
        "db_pemasok": _jload(DB_PEMASOK_FILE),
        "dlm_templates": _jload(DLM_TEMPLATES_FILE),
        "sdl_out_cols": SDL_OUT_COLS,
        "sdl_trx_code_choices": SDL_TRX_CODE_CHOICES,
        "sdl_trx_code_labels": SDL_TRX_CODE_LABELS,
        "sdl_buyer_id_choices": SDL_BUYER_ID_CHOICES,
        "sdl_buyer_id_labels": SDL_BUYER_ID_LABELS,
        "sdl_good_service_choices": SDL_GOOD_SERVICE_CHOICES,
        "sdl_good_service_labels": SDL_GOOD_SERVICE_LABELS,
        "state": public_state(st),
    })


# ---- Upload / Mapping / Process / Reset ---------------------------------------------
@app.route("/api/upload", methods=["POST"])
def upload():
    st=get_state()
    f=request.files.get("file")
    if not f or not f.filename:
        return jsonify({"error":"Tidak ada file."}),400
    try:
        headers,rows=read_raw_excel(f.read(),f.filename)
    except Exception as e:
        return jsonify({"error":str(e)}),400
    st["raw_headers"]=headers; st["raw_rows"]=rows; st["mapping"]={}
    st["src_filename"]=f.filename
    st["processed"]=[]; st["dup_map"]={}; st["faktur_rows"]=[]; st["totals"]={}
    return jsonify({"state":public_state(st)})

@app.route("/api/mapping", methods=["POST"])
def set_mapping():
    st=get_state()
    data=request.get_json(force=True) or {}
    mapping=data.get("mapping",{})
    try:
        st["mapping"]={k:int(v) for k,v in mapping.items()}
    except (TypeError,ValueError):
        return jsonify({"error":"Mapping tidak valid."}),400
    return jsonify({"mapping":st["mapping"]})

@app.route("/api/process", methods=["POST"])
def process_data():
    st=get_state()
    missing=REQUIRED_FIELDS-set(st["mapping"])
    if missing:
        return jsonify({"error":"Field belum lengkap: "+", ".join(sorted(missing))}),400
    try:
        data,dup=apply_mapping_and_process(st["raw_headers"],st["raw_rows"],st["mapping"])
    except Exception as e:
        return jsonify({"error":str(e)}),400
    faktur_rows=build_faktur_rows(data)
    if len(faktur_rows)>TRIAL_INVOICE_LIMIT:
        return trial_upgrade_response(len(faktur_rows), "invoice Pajak Keluaran")
    st["processed"]=data; st["dup_map"]=dup
    st["faktur_rows"]=faktur_rows
    st["totals"]=calc_totals(data)
    apply_buyer_discounts_to_active_data(st,_jload(DB_PEMBELI_FILE))
    return jsonify({"state":public_state(st)})

@app.route("/api/reset", methods=["POST"])
def reset():
    sid=session.get("sid")
    if not sid:
        sid=uuid.uuid4().hex; session["sid"]=sid
    SESSIONS[sid]=default_state()
    return jsonify({"state":public_state(SESSIONS[sid])})


# ---- Templates ------------------------------------------------------------------
@app.route("/api/templates", methods=["GET"])
def list_templates():
    return jsonify({"templates":_jload(TEMPLATES_FILE)})

@app.route("/api/templates", methods=["POST"])
def save_template():
    st=get_state()
    data=request.get_json(force=True) or {}
    name=(data.get("name") or "").strip()
    if not name: return jsonify({"error":"Nama template kosong."}),400
    if not st["mapping"]: return jsonify({"error":"Assign field dulu sebelum menyimpan template."}),400
    templates=_jload(TEMPLATES_FILE)
    templates[name]={"mapping":dict(st["mapping"]),"headers":st["raw_headers"]}
    _jsave(TEMPLATES_FILE,templates)
    return jsonify({"templates":templates})

@app.route("/api/templates/<path:name>", methods=["DELETE"])
def delete_template(name):
    templates=_jload(TEMPLATES_FILE)
    if name in templates:
        del templates[name]; _jsave(TEMPLATES_FILE,templates)
    return jsonify({"templates":templates})

@app.route("/api/templates/<path:name>/apply", methods=["POST"])
def apply_template(name):
    st=get_state()
    templates=_jload(TEMPLATES_FILE)
    tpl=templates.get(name)
    if not tpl: return jsonify({"error":"Template tidak ditemukan."}),404
    try:
        st["mapping"]={k:int(v) for k,v in tpl.get("mapping",{}).items()}
    except (TypeError,ValueError):
        return jsonify({"error":"Mapping pada template tidak valid."}),400
    can_auto=REQUIRED_FIELDS.issubset(st["mapping"]) and bool(st["raw_rows"])
    return jsonify({"mapping":st["mapping"],"can_auto_process":can_auto,"state":public_state(st)})


# ---- Database: Penjual / Pembeli --------------------------------------------------
@app.route("/api/db", methods=["GET"])
def get_db():
    return jsonify({"db_penjual":_jload(DB_PENJUAL_FILE),"db_pembeli":_jload(DB_PEMBELI_FILE)})

@app.route("/api/db/penjual", methods=["POST"])
def save_penjual():
    data=request.get_json(force=True) or {}
    name=(data.get("nama") or "").strip()
    id_tku=(data.get("id_tku") or "").strip()
    npwp=(data.get("npwp") or "").strip()
    if not name: return jsonify({"error":"Isi nama penjual."}),400
    if id_tku and len(id_tku)!=22: return jsonify({"error":"ID TKU harus 22 digit."}),400
    if npwp and len(npwp)!=16: return jsonify({"error":"NPWP harus 16 digit."}),400
    db=_jload(DB_PENJUAL_FILE)
    db[name]={"id_tku":id_tku,"npwp":npwp}
    _jsave(DB_PENJUAL_FILE,db)
    return jsonify({"db_penjual":db})

@app.route("/api/db/penjual/<path:name>", methods=["DELETE"])
def delete_penjual(name):
    db=_jload(DB_PENJUAL_FILE)
    if name in db:
        del db[name]; _jsave(DB_PENJUAL_FILE,db)
    return jsonify({"db_penjual":db})

@app.route("/api/db/pembeli", methods=["POST"])
def save_pembeli():
    st=get_state()
    data=request.get_json(force=True) or {}
    name=(data.get("nama") or "").strip()
    npwp=(data.get("npwp") or "").strip()
    no_dok=(data.get("no_dok") or "").strip()
    diskon_raw=str(data.get("diskon_pct") or "").replace("%","").strip()
    alamat=(data.get("alamat") or "").strip()
    if not name: return jsonify({"error":"Isi nama pembeli."}),400
    if npwp and len(npwp)!=16: return jsonify({"error":"NPWP harus 16 digit."}),400
    if no_dok and len(no_dok)!=22: return jsonify({"error":"No. Dok harus 22 digit."}),400
    if diskon_raw and not re.match(r"^\d+([,.]\d+)?$",diskon_raw):
        return jsonify({"error":"Diskon harus angka persen, contoh 5 atau 4,25. Kosongkan jika tidak ada diskon."}),400
    db=_jload(DB_PEMBELI_FILE)
    db[name]={"nama":name,"npwp":npwp,"no_dok":no_dok,
              "diskon_pct":fmt_pct(parse_num(diskon_raw)) if diskon_raw else "",
              "alamat":alamat}
    _jsave(DB_PEMBELI_FILE,db)
    changed=apply_buyer_discounts_to_active_data(st,db,name)
    return jsonify({"db_pembeli":db,"changed":changed,"state":public_state(st)})

@app.route("/api/db/pembeli/<path:name>", methods=["DELETE"])
def delete_pembeli(name):
    db=_jload(DB_PEMBELI_FILE)
    if name in db:
        del db[name]; _jsave(DB_PEMBELI_FILE,db)
    return jsonify({"db_pembeli":db})

@app.route("/api/db/apply_discounts", methods=["POST"])
def apply_all_discounts():
    st=get_state()
    db_pembeli=_jload(DB_PEMBELI_FILE)
    changed=apply_buyer_discounts_to_active_data(st,db_pembeli)
    return jsonify({"changed":changed,"state":public_state(st)})

@app.route("/api/db/pemasok", methods=["POST"])
def save_pemasok():
    data=request.get_json(force=True) or {}
    name=(data.get("nama") or "").strip()
    npwp=(data.get("npwp") or "").strip()
    if not name: return jsonify({"error":"Isi nama pemasok."}),400
    if npwp and len(npwp)!=16: return jsonify({"error":"NPWP harus 16 digit."}),400
    db=_jload(DB_PEMASOK_FILE)
    db[name]={"nama":name,"npwp":npwp}
    _jsave(DB_PEMASOK_FILE,db)
    return jsonify({"db_pemasok":db})

@app.route("/api/db/pemasok/<path:name>", methods=["DELETE"])
def delete_pemasok(name):
    db=_jload(DB_PEMASOK_FILE)
    if name in db:
        del db[name]; _jsave(DB_PEMASOK_FILE,db)
    return jsonify({"db_pemasok":db})


# ---- Faktur ---------------------------------------------------------------------
@app.route("/api/faktur/bulk_apply", methods=["POST"])
def faktur_bulk_apply():
    st=get_state()
    data=request.get_json(force=True) or {}
    mode=data.get("mode"); scope=data.get("scope"); name=(data.get("name") or "").strip()
    indices=data.get("indices",[]) or []
    if mode not in ("penjual","pembeli"):
        return jsonify({"error":"Mode tidak dikenal."}),400
    if not name:
        return jsonify({"error":f"Pilih nama {'penjual' if mode=='penjual' else 'pembeli'} dari dropdown."}),400
    db_p=_jload(DB_PENJUAL_FILE); db_b=_jload(DB_PEMBELI_FILE)
    if mode=="penjual":
        d=db_p.get(name,{}); id_tku=d.get("id_tku","")
    else:
        d=db_b.get(name,{})
    if scope=="all":
        indices=list(range(len(st["faktur_rows"])))
    else:
        indices=[i for i in indices if isinstance(i,int) and 0<=i<len(st["faktur_rows"])]
    if not indices:
        return jsonify({"error":"Pilih baris dulu (klik / Ctrl+klik)."}),400
    for i in indices:
        r=st["faktur_rows"][i]
        if mode=="penjual":
            r["id_tku_penjual"]=id_tku
        else:
            r["npwp_pembeli"]=d.get("npwp","")
            r["no_dok_pembeli"]=d.get("no_dok","")
            r["nama_pembeli"]=d.get("nama","")
            r["alamat_pembeli"]=d.get("alamat","")
            r["id_tku_pembeli"]=d.get("no_dok","")
            r["diskon_rate"]=parse_num(d.get("diskon_pct",""))
    if mode=="pembeli":
        refs=[st["faktur_rows"][i]["referensi"] for i in indices]
        recalc_detail_for_refs(st,refs,d.get("diskon_pct",""))
    label="semua baris" if scope=="all" else f"{len(indices)} baris"
    return jsonify({"state":public_state(st),"label":label,"name":name})

@app.route("/api/faktur/date_bulk_apply", methods=["POST"])
def faktur_date_bulk_apply():
    st=get_state()
    data=request.get_json(force=True) or {}
    tgl=_normalize_invoice_date(data.get("tgl_faktur"))
    if not tgl:
        return jsonify({"error":"Tanggal Faktur tidak valid. Gunakan DD/MM/YYYY, YYYY-MM-DD, atau format seperti 10 Jun 2026."}),400
    scope=data.get("scope")
    indices=data.get("indices",[]) or []
    if scope=="all":
        indices=list(range(len(st["faktur_rows"])))
    elif scope=="problem":
        indices=collect_faktur_date_diagnostics(st["faktur_rows"])["problem_indices"]
    else:
        indices=[i for i in indices if isinstance(i,int) and 0<=i<len(st["faktur_rows"])]
    if not indices:
        msg="Tidak ada tanggal bermasalah." if scope=="problem" else "Pilih baris faktur dulu (klik / Ctrl+klik)."
        return jsonify({"error":msg}),400
    refs=[]
    for i in indices:
        r=st["faktur_rows"][i]
        r["tgl_faktur"]=tgl
        refs.append(r["referensi"])
    apply_invoice_date_to_refs(st,refs,tgl)
    label="semua faktur" if scope=="all" else ("tanggal bermasalah" if scope=="problem" else f"{len(indices)} faktur")
    return jsonify({"state":public_state(st),"label":label,"tgl_faktur":tgl})

@app.route("/api/faktur/<int:idx>", methods=["POST"])
def faktur_update(idx):
    st=get_state()
    if idx<0 or idx>=len(st["faktur_rows"]):
        return jsonify({"error":"Baris tidak ditemukan."}),404
    data=request.get_json(force=True) or {}
    tgl=None
    if "tgl_faktur" in data:
        tgl=_normalize_invoice_date(data.get("tgl_faktur"))
        if not tgl:
            return jsonify({"error":"Tanggal Faktur tidak valid. Gunakan DD/MM/YYYY, YYYY-MM-DD, atau format seperti 10 Jun 2026."}),400
    id_tku=(data.get("id_tku_penjual") or "").strip()
    diskon_txt=str(data.get("diskon_rate") or "").replace("%","").strip()
    npwp=(data.get("npwp_pembeli") or "").strip()
    no_dok=(data.get("no_dok_pembeli") or "").strip()
    nama=(data.get("nama_pembeli") or "").strip()
    alamat=(data.get("alamat_pembeli") or "").strip()
    warns=[]
    if id_tku and len(id_tku)!=22: warns.append(f"ID TKU harus 22 digit ({len(id_tku)} saat ini)")
    if diskon_txt and not re.match(r"^\d+([,.]\d+)?$",diskon_txt): warns.append("Diskon harus angka persen, contoh 5 atau 4,25")
    if npwp and len(npwp)!=16: warns.append(f"NPWP harus 16 digit ({len(npwp)} saat ini)")
    if no_dok and len(no_dok)!=22: warns.append(f"No. Dok harus 22 digit ({len(no_dok)} saat ini)")
    if warns and not data.get("force"):
        return jsonify({"warnings":warns})
    r=st["faktur_rows"][idx]
    if tgl:
        r["tgl_faktur"]=tgl
        apply_invoice_date_to_refs(st,[r["referensi"]],tgl)
    r["id_tku_penjual"]=id_tku; r["npwp_pembeli"]=npwp; r["no_dok_pembeli"]=no_dok
    r["nama_pembeli"]=nama; r["alamat_pembeli"]=alamat
    r["id_tku_pembeli"]=no_dok
    r["diskon_rate"]=parse_num(diskon_txt)
    recalc_detail_for_refs(st,[r["referensi"]],r.get("diskon_rate",0))
    return jsonify({"state":public_state(st)})


# ---- Detail ---------------------------------------------------------------------
@app.route("/api/detail/apply_defaults", methods=["POST"])
def detail_apply_defaults():
    st=get_state()
    data=request.get_json(force=True) or {}
    opt=_choice_code(data.get("opt"),"A")
    unit=_choice_code(data.get("unit"),"UM.0021")
    code=(data.get("code") or "").strip() or "000000"
    scope=data.get("scope")
    if scope=="all":
        indices=list(range(len(st["processed"])))
    else:
        indices=[i for i in (data.get("indices") or []) if isinstance(i,int) and 0<=i<len(st["processed"])]
        if not indices:
            return jsonify({"error":"Pilih baris detail dulu (klik / Ctrl+klik)."}),400
    for i in indices:
        st["processed"][i]["detail_opt"]=opt
        st["processed"][i]["detail_code"]=code
        st["processed"][i]["detail_unit"]=unit
    label="semua detail" if scope=="all" else f"{len(indices)} detail"
    return jsonify({"state":public_state(st),"label":label,
                     "opt_label":_detail_opt_label(opt),"unit":unit})


# ---- Export ---------------------------------------------------------------------
@app.route("/api/export/warnings")
def export_warnings():
    st=get_state()
    if not st["processed"]:
        return jsonify({"error":"Proses data dulu."}),400
    return jsonify({"blockers":collect_export_blockers(st["faktur_rows"]),
                    "warnings":collect_export_warnings(st["faktur_rows"],st["processed"])})

@app.route("/api/export/xlsx")
def export_xlsx():
    st=get_state()
    if not st["processed"]:
        return jsonify({"error":"Proses data dulu."}),400
    blockers=collect_export_blockers(st["faktur_rows"])
    if blockers:
        return jsonify({"error":"Lengkapi data penjual dan pembeli sebelum export.",
                        "blockers":blockers}),400
    wb=openpyxl.Workbook()
    ws1=wb.active; ws1.title="PajakKeluaran"
    export_tpk_sheet(ws1,st["processed"],st["dup_map"])
    ws2=wb.create_sheet("Faktur")
    npwp_p=(st["faktur_rows"][0].get("id_tku_penjual","")[:16] if st["faktur_rows"] else "")
    export_faktur_sheet(ws2,st["faktur_rows"],npwp_p)
    ws3=wb.create_sheet("DetailFaktur")
    export_detail_faktur_sheet(ws3,st["processed"],st["faktur_rows"])
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    fname=f"{_safe_stem(st['src_filename'])}_cortex.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/export/xml")
def export_xml():
    st=get_state()
    if not st["processed"]:
        return jsonify({"error":"Proses data dulu."}),400
    blockers=collect_export_blockers(st["faktur_rows"])
    if blockers:
        return jsonify({"error":"Lengkapi data penjual dan pembeli sebelum export.",
                        "blockers":blockers}),400
    xml_str=generate_coretax_xml(st["processed"],st["faktur_rows"])
    data_bytes=xml_str.encode("utf-8-sig")
    fname=f"{_safe_stem(st['src_filename'])}_coretax.xml"
    _record_export_history(st)
    return send_file(io.BytesIO(data_bytes), as_attachment=True, download_name=fname, mimetype="application/xml")


# ---- Doc Lain Masukan (input special documents) -----------------------------------
@app.route("/api/dlm/upload", methods=["POST"])
def dlm_upload():
    st=get_state()
    f=request.files.get("file")
    if not f or not f.filename:
        return jsonify({"error":"Tidak ada file."}),400
    try:
        headers,rows=read_raw_excel(f.read(),f.filename)
    except Exception as e:
        return jsonify({"error":str(e)}),400
    dlm=st["dlm"]
    dlm["raw_headers"]=headers; dlm["raw_rows"]=rows; dlm["mapping"]={}
    dlm["src_filename"]=f.filename
    dlm["processed"]=[]; dlm["totals"]={}
    return jsonify({"state":public_state(st)})

@app.route("/api/dlm/mapping", methods=["POST"])
def dlm_set_mapping():
    st=get_state()
    data=request.get_json(force=True) or {}
    mapping=data.get("mapping",{})
    try:
        st["dlm"]["mapping"]={k:int(v) for k,v in mapping.items()}
    except (TypeError,ValueError):
        return jsonify({"error":"Mapping tidak valid."}),400
    return jsonify({"mapping":st["dlm"]["mapping"]})

@app.route("/api/dlm/process", methods=["POST"])
def dlm_process_data():
    st=get_state()
    dlm=st["dlm"]
    missing=DLM_REQUIRED_FIELDS-set(dlm["mapping"])
    if missing:
        return jsonify({"error":"Field belum lengkap: "+", ".join(sorted(missing))}),400
    try:
        data=apply_dlm_mapping_and_process(dlm["raw_headers"],dlm["raw_rows"],dlm["mapping"])
    except Exception as e:
        return jsonify({"error":str(e)}),400
    if len(data)>TRIAL_INVOICE_LIMIT:
        return trial_upgrade_response(len(data), "dokumen Doc Lain Masukan")
    dlm["processed"]=data
    dlm["totals"]=calc_dlm_totals(data)
    return jsonify({"state":public_state(st)})

@app.route("/api/dlm/reset", methods=["POST"])
def dlm_reset():
    st=get_state()
    st["dlm"]=default_dlm_state()
    return jsonify({"state":public_state(st)})

@app.route("/api/dlm/templates", methods=["GET"])
def dlm_list_templates():
    return jsonify({"templates":_jload(DLM_TEMPLATES_FILE)})

@app.route("/api/dlm/templates", methods=["POST"])
def dlm_save_template():
    st=get_state(); dlm=st["dlm"]
    data=request.get_json(force=True) or {}
    name=(data.get("name") or "").strip()
    if not name: return jsonify({"error":"Nama template kosong."}),400
    if not dlm["mapping"]: return jsonify({"error":"Assign field dulu sebelum menyimpan template."}),400
    templates=_jload(DLM_TEMPLATES_FILE)
    templates[name]={"mapping":dict(dlm["mapping"]),"headers":dlm["raw_headers"]}
    _jsave(DLM_TEMPLATES_FILE,templates)
    return jsonify({"templates":templates})

@app.route("/api/dlm/templates/<path:name>", methods=["DELETE"])
def dlm_delete_template(name):
    templates=_jload(DLM_TEMPLATES_FILE)
    if name in templates:
        del templates[name]; _jsave(DLM_TEMPLATES_FILE,templates)
    return jsonify({"templates":templates})

@app.route("/api/dlm/templates/<path:name>/apply", methods=["POST"])
def dlm_apply_template(name):
    st=get_state(); dlm=st["dlm"]
    templates=_jload(DLM_TEMPLATES_FILE)
    tpl=templates.get(name)
    if not tpl: return jsonify({"error":"Template tidak ditemukan."}),404
    try:
        dlm["mapping"]={k:int(v) for k,v in tpl.get("mapping",{}).items()}
    except (TypeError,ValueError):
        return jsonify({"error":"Mapping pada template tidak valid."}),400
    can_auto=DLM_REQUIRED_FIELDS.issubset(dlm["mapping"]) and bool(dlm["raw_rows"])
    return jsonify({"mapping":dlm["mapping"],"can_auto_process":can_auto,"state":public_state(st)})


# ---- SPT Dokumen Lain (pajak keluaran digunggung / RetailInvoiceBulk) -------------
def _sdl_row_from_payload(data,base=None):
    """Validasi payload baris SDL; return (row,None) atau (None,error)."""
    r=dict(base) if base else {
        "serial_no":"","doc_date":"","trx_code":"Normal","buyer_id_opt":"NPWP",
        "buyer_id_number":"","good_service_opt":"A","buyer_name":"","info":"0",
        "vat":0.0,"stlg":0.0,
    }
    buyer_id_number=(data.get("buyer_id_number") or "").strip()
    if buyer_id_number and len(buyer_id_number)!=16:
        return None,f"NPWP/NIK Pembeli harus 16 digit ({len(buyer_id_number)} saat ini)."
    trx_code=_choice_code(data.get("trx_code"),r["trx_code"])
    buyer_id_opt=_choice_code(data.get("buyer_id_opt"),r["buyer_id_opt"])
    good_service_opt=_choice_code(data.get("good_service_opt"),r["good_service_opt"])
    if trx_code not in SDL_TRX_CODE_LABELS:
        return None,"TrxCode tidak dikenal."
    if buyer_id_opt not in SDL_BUYER_ID_LABELS:
        return None,"BuyerIdOpt tidak dikenal."
    if good_service_opt not in SDL_GOOD_SERVICE_LABELS:
        return None,"GoodServiceOpt tidak dikenal."
    vat_txt=str(data.get("vat") if data.get("vat") is not None else "").strip()
    if vat_txt:
        r["vat"]=round2(parse_num(vat_txt))
    if r["vat"]<0:
        return None,"Nominal PPN (VAT) tidak valid."
    doc_date_txt=str(data.get("doc_date") or "").strip()
    if doc_date_txt:
        iso=_date_to_iso(doc_date_txt)
        if not re.match(r"\d{4}-\d{2}-\d{2}$",iso or ""):
            return None,"Tanggal transaksi tidak valid (pakai format YYYY-MM-DD atau DD/MM/YYYY)."
        r["doc_date"]=iso
    if not r["doc_date"]:
        r["doc_date"]=date.today().strftime("%Y-%m-%d")
    r["trx_code"]=trx_code
    r["buyer_id_opt"]=buyer_id_opt
    r["good_service_opt"]=good_service_opt
    r["buyer_id_number"]=buyer_id_number
    r["buyer_name"]=(data.get("buyer_name") or "").strip() or r["buyer_name"]
    r["serial_no"]=(data.get("serial_no") or "").strip() or r["serial_no"]
    r["info"]=str(data.get("info") or "").strip() or "0"
    apply_sdl_calculation(r)
    stlg_txt=str(data.get("stlg") or "").strip()
    r["stlg"]=round2(parse_num(stlg_txt)) if stlg_txt else 0.0
    return r,None

def _sdl_renumber(sdl):
    for i,r in enumerate(sdl["processed"],1):
        r["no"]=i
    sdl["totals"]=calc_sdl_totals(sdl["processed"])

@app.route("/api/sdl/reset", methods=["POST"])
def sdl_reset():
    st=get_state()
    st["sdl"]=default_sdl_state()
    return jsonify({"state":public_state(st)})

@app.route("/api/sdl/rows", methods=["POST"])
def sdl_add_row():
    st=get_state()
    sdl=st["sdl"]
    if len(sdl["processed"])>=TRIAL_INVOICE_LIMIT:
        return trial_upgrade_response(len(sdl["processed"])+1, "dokumen SPT")
    data=request.get_json(force=True) or {}
    missing=[]
    if not str(data.get("vat") or "").strip(): missing.append("PPN (VAT)")
    if not str(data.get("buyer_name") or "").strip(): missing.append("Pembeli")
    if not str(data.get("serial_no") or "").strip(): missing.append("No. Seri")
    if not str(data.get("doc_date") or "").strip(): missing.append("Tgl")
    if missing:
        return jsonify({"error":"Wajib diisi dulu: "+", ".join(missing)+"."}),400
    r,err=_sdl_row_from_payload(data)
    if err: return jsonify({"error":err}),400
    sdl["processed"].append(r)
    _sdl_renumber(sdl)
    if not sdl.get("tax_period_month"):
        m=re.match(r"(\d{4})-(\d{2})-\d{2}$",r["doc_date"])
        if m:
            sdl["tax_period_month"]=int(m.group(2))
            sdl["tax_period_year"]=int(m.group(1))
    return jsonify({"state":public_state(st)})

@app.route("/api/sdl/row/<int:idx>", methods=["POST"])
def sdl_row_update(idx):
    st=get_state()
    sdl=st["sdl"]
    if idx<0 or idx>=len(sdl["processed"]):
        return jsonify({"error":"Baris tidak ditemukan."}),404
    r,err=_sdl_row_from_payload(request.get_json(force=True) or {},base=sdl["processed"][idx])
    if err: return jsonify({"error":err}),400
    sdl["processed"][idx]=r
    _sdl_renumber(sdl)
    return jsonify({"state":public_state(st)})

@app.route("/api/sdl/row/<int:idx>", methods=["DELETE"])
def sdl_delete_row(idx):
    st=get_state()
    sdl=st["sdl"]
    if idx<0 or idx>=len(sdl["processed"]):
        return jsonify({"error":"Baris tidak ditemukan."}),404
    del sdl["processed"][idx]
    _sdl_renumber(sdl)
    return jsonify({"state":public_state(st)})

@app.route("/api/sdl/bulk_apply", methods=["POST"])
def sdl_bulk_apply():
    st=get_state()
    sdl=st["sdl"]
    data=request.get_json(force=True) or {}
    name=(data.get("name") or "").strip()
    scope=data.get("scope")
    indices=data.get("indices",[]) or []
    if not name:
        return jsonify({"error":"Pilih nama pembeli dari dropdown."}),400
    db=_jload(DB_PEMBELI_FILE)
    d=db.get(name,{})
    if scope=="all":
        indices=list(range(len(sdl["processed"])))
    else:
        indices=[i for i in indices if isinstance(i,int) and 0<=i<len(sdl["processed"])]
    if not indices:
        return jsonify({"error":"Pilih baris dulu (klik / Ctrl+klik)."}),400
    for i in indices:
        r=sdl["processed"][i]
        r["buyer_name"]=d.get("nama",name)
        r["buyer_id_number"]=d.get("npwp","")
    label="semua baris" if scope=="all" else f"{len(indices)} baris"
    return jsonify({"state":public_state(st),"label":label,"name":name})

@app.route("/api/sdl/defaults", methods=["POST"])
def sdl_defaults():
    st=get_state()
    sdl=st["sdl"]
    data=request.get_json(force=True) or {}
    trx_code=_choice_code(data.get("trx_code"),"Normal")
    buyer_id_opt=_choice_code(data.get("buyer_id_opt"),"NPWP")
    good_service_opt=_choice_code(data.get("good_service_opt"),"A")
    scope=data.get("scope")
    indices=data.get("indices",[]) or []
    if trx_code not in SDL_TRX_CODE_LABELS:
        return jsonify({"error":"TrxCode tidak dikenal."}),400
    if buyer_id_opt not in SDL_BUYER_ID_LABELS:
        return jsonify({"error":"BuyerIdOpt tidak dikenal."}),400
    if good_service_opt not in SDL_GOOD_SERVICE_LABELS:
        return jsonify({"error":"GoodServiceOpt tidak dikenal."}),400
    if scope=="all":
        indices=list(range(len(sdl["processed"])))
    else:
        indices=[i for i in indices if isinstance(i,int) and 0<=i<len(sdl["processed"])]
    if not indices:
        return jsonify({"error":"Pilih baris dulu (klik / Ctrl+klik)."}),400
    for i in indices:
        r=sdl["processed"][i]
        r["trx_code"]=trx_code
        r["buyer_id_opt"]=buyer_id_opt
        r["good_service_opt"]=good_service_opt
    label="semua baris" if scope=="all" else f"{len(indices)} baris"
    return jsonify({"state":public_state(st),"label":label,
                    "trx_code":trx_code,"buyer_id_opt":buyer_id_opt,"good_service_opt":good_service_opt})

@app.route("/api/sdl/company", methods=["POST"])
def sdl_set_company():
    st=get_state()
    sdl=st["sdl"]
    data=request.get_json(force=True) or {}
    npwp=(data.get("npwp") or "").strip()
    if npwp and len(npwp)!=16:
        return jsonify({"error":"NPWP harus 16 digit."}),400
    try:
        month=int(data.get("month") or 0)
        year=int(data.get("year") or 0)
    except (TypeError,ValueError):
        return jsonify({"error":"Masa/Tahun Pajak tidak valid."}),400
    if month and not 1<=month<=12:
        return jsonify({"error":"Masa Pajak harus 1-12."}),400
    if year and not 2000<=year<=2100:
        return jsonify({"error":"Tahun Pajak tidak valid."}),400
    sdl["company_npwp"]=npwp
    sdl["tax_period_month"]=month
    sdl["tax_period_year"]=year
    return jsonify({"state":public_state(st)})

@app.route("/api/sdl/export/warnings")
def sdl_export_warnings():
    st=get_state(); sdl=st["sdl"]
    if not sdl["processed"]:
        return jsonify({"error":"Proses data dulu."}),400
    return jsonify({"warnings":collect_sdl_export_warnings(sdl)})

@app.route("/api/sdl/export/xml")
def sdl_export_xml():
    st=get_state(); sdl=st["sdl"]
    if not sdl["processed"]:
        return jsonify({"error":"Isi data dulu."}),400
    n=sum(1 for r in sdl["processed"]
          if not str(r.get("buyer_name","")).strip()
          or not str(r.get("serial_no","")).strip()
          or not str(r.get("doc_date","")).strip())
    if n:
        return jsonify({"error":f"PPN, Pembeli, No. Seri, dan Tgl wajib diisi di semua baris sebelum export ({n} baris belum lengkap)."}),400
    month=int(sdl.get("tax_period_month") or 0)
    year=int(sdl.get("tax_period_year") or 0)
    xml_str=generate_sdl_xml(sdl["processed"],sdl.get("company_npwp",""),month,year)
    data_bytes=xml_str.encode("utf-8-sig")
    fname=f"SPT_DOKUMEN_LAIN_{month:02d}_{year}.xml" if month and year else "SPT_DOKUMEN_LAIN.xml"
    return send_file(io.BytesIO(data_bytes), as_attachment=True, download_name=fname, mimetype="application/xml")

@app.route("/api/dlm/row/<int:idx>", methods=["POST"])
def dlm_row_update(idx):
    st=get_state()
    dlm=st["dlm"]
    if idx<0 or idx>=len(dlm["processed"]):
        return jsonify({"error":"Baris tidak ditemukan."}),404
    data=request.get_json(force=True) or {}
    seller_tin=(data.get("seller_tin") or "").strip()
    if seller_tin and len(seller_tin)!=16:
        return jsonify({"error":f"NPWP Pemasok harus 16 digit ({len(seller_tin)} saat ini)."}),400
    r=dlm["processed"][idx]
    trx_type=_choice_code(data.get("trx_type"),r["trx_type"])
    trx_code=_choice_code(data.get("trx_code"),r["trx_code"])
    trx_document=_choice_code(data.get("trx_document"),r["trx_document"])
    if trx_type not in DLM_TRX_TYPE_LABELS:
        return jsonify({"error":"TrxType tidak dikenal."}),400
    if trx_code not in DLM_TRX_DETAIL_LABELS:
        return jsonify({"error":"TrxCode tidak dikenal."}),400
    if trx_document not in DLM_TRX_DOC_LABELS:
        return jsonify({"error":"TrxDocument tidak dikenal."}),400
    r["trx_type"]=trx_type
    r["trx_code"]=trx_code
    r["trx_document"]=trx_document
    r["seller_tin"]=seller_tin
    r["seller_name"]=(data.get("seller_name") or "").strip() or r["seller_name"]
    stlg_txt=str(data.get("stlg") or "").strip()
    r["stlg"]=round2(parse_num(stlg_txt)) if stlg_txt else 0.0
    return jsonify({"state":public_state(st)})

@app.route("/api/dlm/bulk_apply", methods=["POST"])
def dlm_bulk_apply():
    st=get_state()
    dlm=st["dlm"]
    data=request.get_json(force=True) or {}
    name=(data.get("name") or "").strip()
    scope=data.get("scope")
    indices=data.get("indices",[]) or []
    if not name:
        return jsonify({"error":"Pilih nama pemasok dari dropdown."}),400
    db=_jload(DB_PEMASOK_FILE)
    d=db.get(name,{})
    if scope=="all":
        indices=list(range(len(dlm["processed"])))
    else:
        indices=[i for i in indices if isinstance(i,int) and 0<=i<len(dlm["processed"])]
    if not indices:
        return jsonify({"error":"Pilih baris dulu (klik / Ctrl+klik)."}),400
    for i in indices:
        r=dlm["processed"][i]
        r["seller_name"]=d.get("nama",name)
        r["seller_tin"]=d.get("npwp","")
    label="semua baris" if scope=="all" else f"{len(indices)} baris"
    return jsonify({"state":public_state(st),"label":label,"name":name})

@app.route("/api/dlm/trx_defaults", methods=["POST"])
def dlm_trx_defaults():
    st=get_state()
    dlm=st["dlm"]
    data=request.get_json(force=True) or {}
    trx_type=_choice_code(data.get("trx_type"),"04")
    trx_code=_choice_code(data.get("trx_code"),"01")
    trx_document=_choice_code(data.get("trx_document"),"8")
    scope=data.get("scope")
    indices=data.get("indices",[]) or []
    if trx_type not in DLM_TRX_TYPE_LABELS:
        return jsonify({"error":"TrxType tidak dikenal."}),400
    if trx_code not in DLM_TRX_DETAIL_LABELS:
        return jsonify({"error":"TrxCode tidak dikenal."}),400
    if trx_document not in DLM_TRX_DOC_LABELS:
        return jsonify({"error":"TrxDocument tidak dikenal."}),400
    if scope=="all":
        indices=list(range(len(dlm["processed"])))
    else:
        indices=[i for i in indices if isinstance(i,int) and 0<=i<len(dlm["processed"])]
    if not indices:
        return jsonify({"error":"Pilih baris dulu (klik / Ctrl+klik)."}),400
    for i in indices:
        r=dlm["processed"][i]
        r["trx_type"]=trx_type
        r["trx_code"]=trx_code
        r["trx_document"]=trx_document
    label="semua baris" if scope=="all" else f"{len(indices)} baris"
    return jsonify({"state":public_state(st),"label":label,
                    "trx_type":trx_type,"trx_code":trx_code,"trx_document":trx_document})

@app.route("/api/dlm/company", methods=["POST"])
def dlm_set_company():
    st=get_state()
    data=request.get_json(force=True) or {}
    npwp=(data.get("npwp") or "").strip()
    if npwp and len(npwp)!=16:
        return jsonify({"error":"NPWP harus 16 digit."}),400
    st["dlm"]["company_npwp"]=npwp
    return jsonify({"state":public_state(st)})

@app.route("/api/dlm/export/warnings")
def dlm_export_warnings():
    st=get_state(); dlm=st["dlm"]
    if not dlm["processed"]:
        return jsonify({"error":"Proses data dulu."}),400
    return jsonify({"warnings":collect_dlm_export_warnings(dlm)})

@app.route("/api/dlm/export/xlsx")
def dlm_export_xlsx():
    st=get_state(); dlm=st["dlm"]
    if not dlm["processed"]:
        return jsonify({"error":"Proses data dulu."}),400
    wb=openpyxl.Workbook()
    ws=wb.active; ws.title="DATA"
    export_dlm_sheet(ws,dlm["processed"],dlm.get("company_npwp",""))
    export_dlm_reference_sheet(wb.create_sheet("Trx Detail"),"Trx Detail",DLM_TRX_DETAIL_LABELS)
    export_dlm_reference_sheet(wb.create_sheet("Trx Doc"),"Trx Doc",DLM_TRX_DOC_LABELS)
    export_dlm_reference_sheet(wb.create_sheet("Trx TYpe"),"Trx TYpe",DLM_TRX_TYPE_LABELS)
    wb.active=0
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    fname=f"{_safe_stem(dlm['src_filename'])}_doclain.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                      mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/api/dlm/export/xml")
def dlm_export_xml():
    st=get_state(); dlm=st["dlm"]
    if not dlm["processed"]:
        return jsonify({"error":"Proses data dulu."}),400
    xml_str=generate_dlm_xml(dlm["processed"],dlm.get("company_npwp",""))
    data_bytes=xml_str.encode("utf-8-sig")
    fname=f"{_safe_stem(dlm['src_filename'])}_doclain.xml"
    return send_file(io.BytesIO(data_bytes), as_attachment=True, download_name=fname, mimetype="application/xml")


# ---- Dashboard ----------------------------------------------------------------------
@app.route("/api/dashboard")
def dashboard_data():
    history=_jload(EXPORT_HISTORY_FILE)
    items=sorted(history.values(),key=lambda x:x.get("exported_at",""),reverse=True)
    return jsonify({"total":len(items),"items":items})

@app.route("/api/dashboard/reset", methods=["POST"])
def dashboard_reset():
    _jsave(EXPORT_HISTORY_FILE,{})
    return jsonify({"total":0,"items":[]})


# ---- Settings ---------------------------------------------------------------------
@app.route("/api/settings/theme", methods=["POST"])
def set_theme():
    data=request.get_json(force=True) or {}
    name=data.get("theme")
    if name not in THEMES:
        return jsonify({"error":"Tema tidak dikenal."}),400
    _jsave(SETTINGS_FILE,{"theme":name})
    return jsonify({"theme":name})


if __name__=="__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="127.0.0.1",port=port,debug=False)
